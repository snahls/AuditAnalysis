import openpyxl
import json
from pathlib import Path
import os
root = os.path.dirname(os.path.realpath(__file__))
print(root)

def extractHeadingSeverity(fullheading):
    temp = fullheading.split()
    heading = ""
    severity = ""
    if temp[0].startswith("ERROR") or temp[0].startswith("WARNING"):
        severity = "Medium"
    elif temp[0].startswith("CRITICAL") or temp[0].startswith("EMERGENCY"):
        severity = "High"
    elif temp[0].startswith("NOTICE"):
        severity = "Low"
    elif temp[0].startswith("INFO"):
        severity = "Info Only"

    for i in range(1, len(temp)):
        heading = heading + " " + temp[i]
    heading = heading.lstrip()
    return severity,heading

def extractSympCondMitAddEvi(information):
    info = {}
    if "SYMPTOMS:" in information:
        a = information.split("SYMPTOMS:")
        b = a[1].split("CONDITIONS:")
        c = b[1].split("MITIGATION:")
        d = c[1].split("ADDITIONAL INFORMATION:")
        e = d[1].split("EVIDENCE:")

        try:
            info["Symptoms"] = b[0]
        except IndexError:
            info["Symptoms"] = ""
        try:
            info["Conditions"] = c[0]
        except IndexError:
            info["Conditions"] = ""
        try:
            info["Mitigation"] = d[0]
        except IndexError:
            info["Mitigation"] = ""
        try:
            info["Additional Information"] = e[0]
        except IndexError:
            info["Additional Information"] = ""
        try:
            info["Evidence"] = e[1]
        except IndexError:
            info["Evidence"] = ""
    else:
        info["Symptoms"] = information
        info["Conditions"] = ""
        info["Mitigation"] = ""
        info["Additional Information"] = ""
        info["Evidence"] = ""

    return info

def showTechExcelJson(showTechExcel):
    stExcel = openpyxl.load_workbook(showTechExcel)
    ste1 = stExcel['Sheet1']
    ste2 = stExcel['Sheet2']
    ste3 = stExcel['Sheet3']
    #columnNames = ["Severity", "Heading", "Description", "Audit Recommendation", "Evidence", "Domain", "UCSM Version"]
    data = []
    x = []
    y = []
    ctr = 1
    #developed modules in the show tech parser script
    for i in range (1,ste1.max_row+1):
        x = extractHeadingSeverity(ste1.cell(row=i,column=1).value)
        y = extractSympCondMitAddEvi(ste1.cell(row=i, column=2).value)

        

        data.append({"ID": ctr, "Domain": ste3['B1'].value,"Severity": x[0], "Heading":x[1],
                     "Description":y["Symptoms"]+y["Conditions"],
                     "Audit Recommendation":y["Mitigation"]+y["Additional Information"],
                     "Evidence":y["Evidence"],
                     
                     "UCSM Version": ste3['B2'].value})
        ctr+=1
    #in development modules from the show tech parser script
    
    for i in range(1,ste2.max_row+1):
        x = extractHeadingSeverity(ste2.cell(row=i, column=1).value)
        y = extractSympCondMitAddEvi(ste2.cell(row=i, column=2).value)

        
        if 'Unique CIMC firmware and hardware' in y["Evidence"]:
            temp_result = y["Evidence"]
            t = temp_result.splitlines()
            #print(t[3])
           
            all_server_detail=[]
            for i in range(3, len(t)):
                server_detail ={}
                item = t[i]
                print(item)
                if '|' in item:
                    it = item.split('|')
                    server_detail['num_server'] = it[0]
                    det = it[1].split('/')
                    server_detail['OS Version']  = det[0]
                    server_detail['Server Type']  = det[1]
                    server_detail['Processor Type']  = det[2]
                    server_detail['CNA Type']  = det[3]
                    all_server_detail.append(server_detail)
            #print(json.dumps(all_server_detail, indent=2))
            
        data.append({"ID":ctr,"Domain": ste3['B1'].value,"Severity": x[0], "Heading": x[1],
                     "Description": y["Symptoms"] + y["Conditions"],
                     "Audit Recommendation": y["Mitigation"] + y["Additional Information"],
                     "Evidence": y["Evidence"],
                     
                     "UCSM Version": ste3['B2'].value})
        ctr += 1
    #print(json.dumps(data, indent=2))
    return all_server_detail, data
    

#printing the result for example
#showTechExcelJson(Path(root)/"UCS_ShowTech_check.xlsx")