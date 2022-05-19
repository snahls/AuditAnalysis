import os
import sys
from openpyxl import Workbook, utils
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting import Rule
from openpyxl.styles import Border, Side, Alignment
from openpyxl.chart import (Reference,Series,BarChart3D,LineChart,PieChart3D,PieChart,BarChart,ScatterChart)
import openpyxl.worksheet.dimensions
from datetime import datetime, date
import jsonCreator
import json
import pymongo
import mail_utility
from webexteamssdk import WebexTeamsAPI

from openpyxl.chart.series import DataPoint



scriptPath = os.path.dirname(os.path.realpath(__file__))


thin_border = Border(left=Side(style='thin'), 
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))

thick_border = Border(left=Side(style='thick'), 
					right=Side(style='thick'), 
					top=Side(style='thick'), 
					bottom=Side(style='thick'))

bottom_thick_border = Border(bottom=Side(style='thick'))
top_thick_border = Border(top=Side(style='thick'))

topBottom_thick_border = Border(left=Side(style='thin'), 
					right=Side(style='thin'), 
					top=Side(style='thick'), 
					bottom=Side(style='thick'))


left_thin_border = Border(left=Side(style='thin'), 
					right=Side(style='thick'), 
					top=Side(style='thick'), 
					bottom=Side(style='thick'))

right_thin_border = Border(left=Side(style='thick'), 
					right=Side(style='thin'), 
					top=Side(style='thick'), 
					bottom=Side(style='thick'))

left_thick_border = Border(left=Side(style='thick'), 
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))

right_thick_border = Border(left=Side(style='thin'), 
					right=Side(style='thick'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))


topLeft_thick_border = Border(left=Side(style='thick'), 
					right=Side(style='thin'), 
					top=Side(style='thick'), 
					bottom=Side(style='thin'))

topRight_thick_border = Border(left=Side(style='thin'), 
					right=Side(style='thick'), 
					top=Side(style='thick'), 
					bottom=Side(style='thin'))

bottomLeft_thick_border = Border(left=Side(style='thick'), 
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thick'))

bottomRight_thick_border = Border(left=Side(style='thin'), 
					right=Side(style='thick'), 
					top=Side(style='thin'), 
					bottom=Side(style='thick'))




def getAuditDetails():
	cpyKey = sys.argv[1]
	auditID = sys.argv[2]
	audit_type = sys.argv[3]
	Customer_Name = sys.argv[4]
	email_id = sys.argv[5]
	return cpyKey,auditID,audit_type,Customer_Name,email_id

def getNetRules(audit_type, table_name, exception_name):
		dbName = "CNA_Visualizer"
		dbClient = pymongo.MongoClient('localhost', 27017)
		db = dbClient[dbName]
		query = {"audit_type" : audit_type, "table_name" : table_name , "exception_name" : exception_name}
		res = db['NetRule_NetAdvice'].find(query)
		result = list(res)
		dbClient.close()
		return result

def getNetRulesv2(cpyKey, audit_id, audit_type, table_name, exception_name):
		dbName = "CNA_Visualizer"
		dbClient = pymongo.MongoClient('localhost', 27017)
		db = dbClient[dbName]
		if("- " in exception_name):
			exception_name = exception_name[exception_name.rfind("- ")+2:]

		query = {"Audit_ID" : audit_id, "Audit_Type" : audit_type, "table_name" : table_name , "exception" : exception_name, "jsonFor" : "netRulenetAdvice"}
		res = db[cpyKey].find(query)
		result = list(res)
		if not result:
			query = {"Audit_ID" : audit_id, "Audit_Type" : audit_type, "table_name" : table_name , "exception" : exception_name+" (Exceptions Only)", "jsonFor" : "netRulenetAdvice"}
			#print(query)
			res = db[str(cpyKey)].find(query)
			result = list(res)
		dbClient.close()
			
		return result

def getSev(col):
	sev = ""
	if(col == "0000BCEB"):
		sev = "Critical"
	elif(col == "00E2231A"):
		sev = "High"
	elif(col == "00FBAB18"):
		sev = "Medium"
	elif(col == "00EED202"):
		sev = "Low"
	elif(col == "006ABF4B"):
		sev = "Informational"

	return sev

def getTableSevData(wb):
	sev_arr = ["Critical" , "High" , "Medium" , "Low" , "Informational"]
	table_dict = {}
	for sev in sev_arr:
		ws = wb[sev]
		for i in range(4, ws.max_row+1):
			if(ws.cell(row = i , column = 1).value in table_dict):
				s = getSev(str(ws.cell(row = i , column = 3).fill.start_color.rgb))
				table_dict[ws.cell(row = i , column = 1).value][s] = table_dict[ws.cell(row = i , column = 1).value][s] + 1
			else:
				table_dict[ws.cell(row = i , column = 1).value] = {"Critical" : 0 , "High" : 0 , "Medium" : 0 , "Low" : 0 , "Informational" : 0}
				s = getSev(str(ws.cell(row = i , column = 3).fill.start_color.rgb))
				table_dict[ws.cell(row = i , column = 1).value][s] = table_dict[ws.cell(row = i , column = 1).value][s] + 1

	return table_dict

def create_overview_sheet(wb, audit_information , all_tables, severity_breakdown):
	table_dict = getTableSevData(wb)
	active_sheet = wb["Overview"]
	active_sheet.row_dimensions[1].height = 35
	active_sheet.row_dimensions[2].height = 35
	active_sheet.row_dimensions[3].height = 35
	active_sheet.row_dimensions[4].height = 35
	active_sheet.row_dimensions[5].height = 35
	active_sheet.row_dimensions[6].height = 35
	active_sheet.row_dimensions[7].height = 35
	active_sheet.row_dimensions[8].height = 35
	active_sheet.row_dimensions[9].height = 35
	active_sheet.row_dimensions[10].height = 35
	active_sheet.row_dimensions[11].height = 35
	active_sheet.row_dimensions[12].height = 35


	active_sheet.column_dimensions['A'].width = 45
	active_sheet.column_dimensions['B'].width = 45
	active_sheet.column_dimensions['C'].width = 45
	active_sheet.column_dimensions['D'].width = 45
	active_sheet.column_dimensions['E'].width = 45
	active_sheet.column_dimensions['F'].width = 45
	active_sheet.column_dimensions['G'].width = 45

	active_sheet.cell(row= 1 , column = 1).value = "Audit Information"
	active_sheet.cell(row= 1 , column = 1).border = thick_border
	active_sheet.cell(row= 1 , column = 1).font = Font(bold=True,size='16', color = "000080")
	active_sheet.cell(row= 1 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


	active_sheet.cell(row= 2 , column = 1).value = "Customer Name"
	active_sheet.cell(row= 2 , column = 1).border = topLeft_thick_border
	active_sheet.cell(row= 2 , column = 1).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 2 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 2 , column = 2).value = audit_information[0]["Customer_Name"]
	active_sheet.cell(row= 2 , column = 2).border = topRight_thick_border
	active_sheet.cell(row= 2 , column = 2).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 2 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	
	active_sheet.cell(row= 3 , column = 1).value = "Audit Start Date"
	active_sheet.cell(row= 3 , column = 1).border = left_thick_border
	active_sheet.cell(row= 3 , column = 1).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 3 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 3 , column = 2).value = audit_information[0]["Collection Start Time"]
	active_sheet.cell(row= 3 , column = 2).border = right_thick_border
	active_sheet.cell(row= 3 , column = 2).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 3 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 4 , column = 1).value = "Audit End Date"
	active_sheet.cell(row= 4 , column = 1).border = left_thick_border
	active_sheet.cell(row= 4 , column = 1).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 4 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 4 , column = 2).value = audit_information[0]["Collection End Time"]
	active_sheet.cell(row= 4 , column = 2).border = right_thick_border
	active_sheet.cell(row= 4 , column = 2).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 4 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 5 , column = 1).value = "Audit Time"
	active_sheet.cell(row= 5 , column = 1).border = bottomLeft_thick_border
	active_sheet.cell(row= 5 , column = 1).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 5 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 5 , column = 2).value = audit_information[0]["Collection Period (Days)"]
	active_sheet.cell(row= 5 , column = 2).border = bottomRight_thick_border
	active_sheet.cell(row= 5 , column = 2).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 5 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 1 , column = 5).value = "Device Information"
	active_sheet.cell(row= 1 , column = 5).font = Font(bold=True,size='16', color = "000080")
	active_sheet.cell(row= 1 , column = 5).border = thick_border
	active_sheet.cell(row= 1 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 2 , column = 5).value = "Devices Attempted"
	active_sheet.cell(row= 2 , column = 5).border = topLeft_thick_border
	active_sheet.cell(row= 2 , column = 5).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 2 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 2 , column = 6).value = audit_information[0]["Devices Attempted"]
	active_sheet.cell(row= 2 , column = 6).border = topRight_thick_border
	active_sheet.cell(row= 2 , column = 6).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 2 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 3 , column = 5).value = "Devices Passed"
	active_sheet.cell(row= 3 , column = 5).border = left_thick_border
	active_sheet.cell(row= 3 , column = 5).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 3 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 3 , column = 6).value = audit_information[0]["Devices Passed"]
	active_sheet.cell(row= 3 , column = 6).border = right_thick_border
	active_sheet.cell(row= 3 , column = 6).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 3 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 4 , column = 5).value = "Audit Failed"
	active_sheet.cell(row= 4 , column = 5).border = left_thick_border
	active_sheet.cell(row= 4 , column = 5).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 4 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 4 , column = 6).value = audit_information[0]["Devices Failed/Excluded"]
	active_sheet.cell(row= 4 , column = 6).border = right_thick_border
	active_sheet.cell(row= 4 , column = 6).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 4 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 5 , column = 5).value = "Audit Duplicated"
	active_sheet.cell(row= 5 , column = 5).border = bottomLeft_thick_border
	active_sheet.cell(row= 5 , column = 5).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= 5 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 5 , column = 6).value = audit_information[0]["Devices Duplicated"]
	active_sheet.cell(row= 5 , column = 6).border = bottomRight_thick_border
	active_sheet.cell(row= 5 , column = 6).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 5 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	for k in range(1,50):
		active_sheet.cell(row = 6, column = k).border = bottom_thick_border

	active_sheet.cell(row= 8 , column = 1).value = "FCCAPS"
	active_sheet.cell(row= 8 , column = 1).border = thick_border
	active_sheet.cell(row= 8 , column = 1).font = Font(bold=True,size='16', color = "000080")
	active_sheet.cell(row= 8 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 8 , column = 2).value = "Fault Management"
	active_sheet.cell(row= 8 , column = 2).border = thick_border
	active_sheet.cell(row= 8 , column = 2).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 2).hyperlink = "#'Fault Management'!A1"

	active_sheet.cell(row= 8 , column = 3).value = "Configuration Management"
	active_sheet.cell(row= 8 , column = 3).border = thick_border
	active_sheet.cell(row= 8 , column = 3).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 3).hyperlink = "#'Configuration Management'!A1"

	active_sheet.cell(row= 8 , column = 4).value = "Capacity Management"
	active_sheet.cell(row= 8 , column = 4).border = thick_border
	active_sheet.cell(row= 8 , column = 4).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 4).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 4).hyperlink = "#'Capacity Management'!A1"

	active_sheet.cell(row= 8 , column = 5).value = "Performance Management"
	active_sheet.cell(row= 8 , column = 5).border = thick_border
	active_sheet.cell(row= 8 , column = 5).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 5).hyperlink = "#'Performance Management'!A1"

	active_sheet.cell(row= 8 , column = 6).value = "Security Management"
	active_sheet.cell(row= 8 , column = 6).border = thick_border
	active_sheet.cell(row= 8 , column = 6).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 6).hyperlink = "#'Security Management'!A1"

	active_sheet.cell(row= 8 , column = 7).value = "Verify Manually"
	active_sheet.cell(row= 8 , column = 7).border = thick_border
	active_sheet.cell(row= 8 , column = 7).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 8 , column = 7).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 8 , column = 7).hyperlink = "#'Verify Manually'!A1"

	active_sheet.cell(row= 9 , column = 1).value = "Total Exceptions"
	active_sheet.cell(row= 9 , column = 1).border = thick_border
	active_sheet.cell(row= 9 , column = 1).font = Font(bold=True,size='16', color = "000080")
	active_sheet.cell(row= 9 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	#active_sheet.cell(row= 9 , column = 2).value = int(severity_breakdown[0]["Fault Management"]["Critical"]) + int(severity_breakdown[0]["Fault Management"]["High"]) + int(severity_breakdown[0]["Fault Management"]["Medium"]) + int(severity_breakdown[0]["Fault Management"]["Low"]) + int(severity_breakdown[0]["Fault Management"]["Informational"])
	active_sheet.cell(row= 9 , column = 2).value = wb["Fault Management"].max_row-3
	active_sheet.cell(row= 9 , column = 2).border = thick_border
	active_sheet.cell(row= 9 , column = 2).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


	#active_sheet.cell(row= 9 , column = 3).value = int(severity_breakdown[0]["Configuration Management"]["Critical"]) + int(severity_breakdown[0]["Configuration Management"]["High"]) + int(severity_breakdown[0]["Configuration Management"]["Medium"]) + int(severity_breakdown[0]["Configuration Management"]["Low"]) + int(severity_breakdown[0]["Configuration Management"]["Informational"])
	active_sheet.cell(row= 9 , column = 3).value = wb["Configuration Management"].max_row-3
	active_sheet.cell(row= 9 , column = 3).border = thick_border
	active_sheet.cell(row= 9 , column = 3).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


	#active_sheet.cell(row= 9 , column = 4).value = int(severity_breakdown[0]["Capacity Management"]["Critical"]) + int(severity_breakdown[0]["Capacity Management"]["High"]) + int(severity_breakdown[0]["Capacity Management"]["Medium"]) + int(severity_breakdown[0]["Capacity Management"]["Low"]) + int(severity_breakdown[0]["Capacity Management"]["Informational"])
	active_sheet.cell(row= 9 , column = 4).value = wb["Capacity Management"].max_row-3
	active_sheet.cell(row= 9 , column = 4).border = thick_border
	active_sheet.cell(row= 9 , column = 4).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 4).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


	#active_sheet.cell(row= 9 , column = 5).value = int(severity_breakdown[0]["Performance Management"]["Critical"]) + int(severity_breakdown[0]["Performance Management"]["High"]) + int(severity_breakdown[0]["Performance Management"]["Medium"]) + int(severity_breakdown[0]["Performance Management"]["Low"]) + int(severity_breakdown[0]["Performance Management"]["Informational"])
	active_sheet.cell(row= 9 , column = 5).value = wb["Performance Management"].max_row-3
	active_sheet.cell(row= 9 , column = 5).border = thick_border
	active_sheet.cell(row= 9 , column = 5).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


	#active_sheet.cell(row= 9 , column = 6).value = int(severity_breakdown[0]["Security Management"]["Critical"]) + int(severity_breakdown[0]["Security Management"]["High"]) + int(severity_breakdown[0]["Security Management"]["Medium"]) + int(severity_breakdown[0]["Security Management"]["Low"]) + int(severity_breakdown[0]["Security Management"]["Informational"])
	active_sheet.cell(row= 9 , column = 6).value = wb["Security Management"].max_row-3
	active_sheet.cell(row= 9 , column = 6).border = thick_border
	active_sheet.cell(row= 9 , column = 6).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	
	#active_sheet.cell(row= 9 , column = 7).value = int(severity_breakdown[0]["Verify Manually"]["Critical"]) + int(severity_breakdown[0]["Verify Manually"]["High"]) + int(severity_breakdown[0]["Verify Manually"]["Medium"]) + int(severity_breakdown[0]["Verify Manually"]["Low"]) + int(severity_breakdown[0]["Verify Manually"]["Informational"])
	active_sheet.cell(row= 9 , column = 7).value = wb["Verify Manually"].max_row-3
	active_sheet.cell(row= 9 , column = 7).border = thick_border
	active_sheet.cell(row= 9 , column = 7).font = Font(bold=True,size='12', color = "000080")
	active_sheet.cell(row= 9 , column = 7).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	for k in range(1,50):
		active_sheet.cell(row = 10, column = k).border = bottom_thick_border

	active_sheet.cell(row= 12 , column = 1).value = "Table Name"
	active_sheet.cell(row= 12 , column = 1).border = thick_border
	active_sheet.cell(row= 12 , column = 1).font = Font(bold=True,size='16', color = "000080")
	active_sheet.cell(row= 12 , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	active_sheet.cell(row= 12 , column = 2).value = "Critical"
	active_sheet.cell(row= 12 , column = 2).border = thick_border
	active_sheet.cell(row= 12 , column = 2).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 12 , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 12 , column = 2).fill = PatternFill(start_color='00BCEB',end_color='00BCEB',fill_type='solid')
	active_sheet.cell(row= 12 , column = 2).hyperlink = "#'Critical'!A1"

	active_sheet.cell(row= 12 , column = 3).value = "High"
	active_sheet.cell(row= 12 , column = 3).border = thick_border
	active_sheet.cell(row= 12 , column = 3).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 12 , column = 3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 12 , column = 3).fill = PatternFill(start_color='E2231A',end_color='E2231A',fill_type='solid')
	active_sheet.cell(row= 12 , column = 3).hyperlink = "#'High'!A1"

	active_sheet.cell(row= 12 , column = 4).value = "Medium"
	active_sheet.cell(row= 12 , column = 4).border = thick_border
	active_sheet.cell(row= 12 , column = 4).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 12 , column = 4).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 12 , column = 4).fill = PatternFill(start_color='FBAB18',end_color='FBAB18',fill_type='solid')
	active_sheet.cell(row= 12 , column = 4).hyperlink = "#'Medium'!A1"

	active_sheet.cell(row= 12 , column = 5).value = "Low"
	active_sheet.cell(row= 12 , column = 5).border = thick_border
	active_sheet.cell(row= 12 , column = 5).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 12 , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 12 , column = 5).fill = PatternFill(start_color='EED202',end_color='EED202',fill_type='solid')
	active_sheet.cell(row= 12 , column = 5).hyperlink = "#'Low'!A1"

	active_sheet.cell(row= 12 , column = 6).value = "Informational"
	active_sheet.cell(row= 12 , column = 6).border = thick_border
	active_sheet.cell(row= 12 , column = 6).font = Font(bold=True,size='16', color = "000080", underline='single')
	active_sheet.cell(row= 12 , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= 12 , column = 6).fill = PatternFill(start_color='6ABF4B',end_color='6ABF4B',fill_type='solid')
	active_sheet.cell(row= 12 , column = 6).hyperlink = "#'Informational'!A1"

	rowNum = 13
	for table in table_dict.keys():

		active_sheet.row_dimensions[rowNum].height = 35
		active_sheet.cell(row= rowNum , column = 1).value = table
		active_sheet.cell(row= rowNum , column = 1).border = thick_border
		active_sheet.cell(row= rowNum , column = 1).font = Font(bold=True,size='14', color = "000080" , underline='single')
		active_sheet.cell(row= rowNum , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
		active_sheet.cell(row= rowNum , column = 1).hyperlink = "#'"+table[0:31].replace(":"," ")+"'!A1"

		active_sheet.cell(row= rowNum , column = 2).value = table_dict[table]["Critical"]
		active_sheet.cell(row= rowNum , column = 2).border = thin_border
		active_sheet.cell(row= rowNum , column = 2).font = Font(bold=True,size='12', color = "000080")
		active_sheet.cell(row= rowNum , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


		active_sheet.cell(row= rowNum , column = 3).value = table_dict[table]["High"]
		active_sheet.cell(row= rowNum , column = 3).border = thin_border
		active_sheet.cell(row= rowNum , column = 3).font = Font(bold=True,size='12', color = "000080")
		active_sheet.cell(row= rowNum , column = 3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


		active_sheet.cell(row= rowNum , column = 4).value = table_dict[table]["Medium"]
		active_sheet.cell(row= rowNum , column = 4).border = thin_border
		active_sheet.cell(row= rowNum , column = 4).font = Font(bold=True,size='12', color = "000080")
		active_sheet.cell(row= rowNum , column = 4).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


		active_sheet.cell(row= rowNum , column = 5).value = table_dict[table]["Low"]
		active_sheet.cell(row= rowNum , column = 5).border = thin_border
		active_sheet.cell(row= rowNum , column = 5).font = Font(bold=True,size='12', color = "000080")
		active_sheet.cell(row= rowNum , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


		active_sheet.cell(row= rowNum , column = 6).value = table_dict[table]["Informational"]
		active_sheet.cell(row= rowNum , column = 6).border = right_thick_border
		active_sheet.cell(row= rowNum , column = 6).font = Font(bold=True,size='12', color = "000080")
		active_sheet.cell(row= rowNum , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


		rowNum = rowNum + 1

	active_sheet.row_dimensions[rowNum].height = 35

	active_sheet.cell(row= rowNum , column = 1).value = 'Total'
	active_sheet.cell(row= rowNum , column = 1).border = thick_border
	active_sheet.cell(row= rowNum , column = 1).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 1).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	colStr = 'B'+str(rowNum-1)
	active_sheet.cell(row= rowNum , column = 2).value = "=SUM(B13:"+colStr+")"
	active_sheet.cell(row= rowNum , column = 2).border = thick_border
	active_sheet.cell(row= rowNum , column = 2).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 2).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= rowNum , column = 2).fill = PatternFill(start_color='00BCEB',end_color='00BCEB',fill_type='solid')

	colStr = 'C'+str(rowNum-1)
	active_sheet.cell(row= rowNum , column = 3).value = "=SUM(C13:"+colStr+")"
	active_sheet.cell(row= rowNum , column = 3).border = thick_border
	active_sheet.cell(row= rowNum , column = 3).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= rowNum , column = 3).fill = PatternFill(start_color='E2231A',end_color='E2231A',fill_type='solid')
	
	colStr = 'D'+str(rowNum-1)
	active_sheet.cell(row= rowNum , column = 4).value = "=SUM(D13:"+colStr+")"
	active_sheet.cell(row= rowNum , column = 4).border = thick_border
	active_sheet.cell(row= rowNum , column = 4).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 4).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= rowNum , column = 4).fill = PatternFill(start_color='FBAB18',end_color='FBAB18',fill_type='solid')

	colStr = 'E'+str(rowNum-1)
	active_sheet.cell(row= rowNum , column = 5).value = "=SUM(E13:"+colStr+")"
	active_sheet.cell(row= rowNum , column = 5).border = thick_border
	active_sheet.cell(row= rowNum , column = 5).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 5).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= rowNum , column = 5).fill = PatternFill(start_color='EED202',end_color='EED202',fill_type='solid')

	colStr = 'F'+str(rowNum-1)
	active_sheet.cell(row= rowNum , column = 6).value = "=SUM(F13:"+colStr+")"
	active_sheet.cell(row= rowNum , column = 6).border = thick_border
	active_sheet.cell(row= rowNum , column = 6).font = Font(bold=True,size='14', color = "000080")
	active_sheet.cell(row= rowNum , column = 6).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
	active_sheet.cell(row= rowNum , column = 6).fill = PatternFill(start_color='6ABF4B',end_color='6ABF4B',fill_type='solid')

	colStr = 'F'+str(rowNum-1)
	active_sheet.merge_cells('C3:D3')
	active_sheet['C3'].value = "Total Exceptions"
	active_sheet['C3'].alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet['C3'].font = Font(name="CiscoSans" , size = 22, color = "FFFFFF", bold=True)
	active_sheet['C3'].fill = PatternFill(start_color='010030',end_color='010030',fill_type='solid')
	active_sheet['C3'].border = topLeft_thick_border
	active_sheet['D3'].border = topRight_thick_border

	active_sheet.merge_cells('C4:D4')
	active_sheet['C4'].value = "=SUM(B13:"+colStr+")"
	active_sheet['C4'].alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet['C4'].font = Font(name="CiscoSans" , size = 22, color = "FFFFFF", bold=True)
	active_sheet['C4'].fill = PatternFill(start_color='010030',end_color='010030',fill_type='solid')
	active_sheet['C4'].border = bottomLeft_thick_border
	active_sheet['D4'].border = bottomRight_thick_border


def styleSheet(active_sheet):
	active_sheet.row_dimensions[3].height = 40
	active_sheet.column_dimensions['A'].width = 25
	active_sheet.column_dimensions['B'].width = 25
	active_sheet.column_dimensions['C'].width = 25
	active_sheet.column_dimensions['D'].width = 40
	active_sheet.column_dimensions['E'].width = 25
	active_sheet.column_dimensions['F'].width = 25
	active_sheet.column_dimensions['G'].width = 80
	active_sheet.column_dimensions['H'].width = 30

	active_sheet.merge_cells('A1:B1')
	active_sheet['A1'].value = "Return to Overview"
	active_sheet['A1'].hyperlink = "#'Overview'!A1"
	active_sheet['A1'].alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet['A1'].font = Font(name="CiscoSans" , size = 18, bold=True)
	active_sheet['A1'].border = thin_border

	active_sheet.merge_cells('A2:B2')
	active_sheet['A2'].value = "Return to Exception List"
	active_sheet['A2'].hyperlink = "#'Exceptions'!A1"
	active_sheet['A2'].alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet['A2'].font = Font(name="CiscoSans" , size = 18, bold=True)
	active_sheet['A2'].border = thin_border



	active_sheet.cell(row = 3, column = 1).value = 'Table Name'
	active_sheet.cell(row = 3, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 1).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 1).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 2).value = 'NMS Area'
	active_sheet.cell(row = 3, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 2).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 2).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 3).value = 'Exception'
	active_sheet.cell(row = 3, column = 3).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 3).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 3).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 4).value = 'Affected Devices'
	active_sheet.cell(row = 3, column = 4).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 4).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 4).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 5).value = 'Net Rule'
	active_sheet.cell(row = 3, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 5).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 5).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 6).value = 'Net Advice'
	active_sheet.cell(row = 3, column = 6).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 6).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 6).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 7).value = 'NCE Comments'
	active_sheet.cell(row = 3, column = 7).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 7).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 7).border = topBottom_thick_border

	active_sheet.cell(row = 3, column = 8).value = 'Additional Info'
	active_sheet.cell(row = 3, column = 8).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 3, column = 8).font = Font(name="CiscoSans" , size = 14, bold=True)
	active_sheet.cell(row = 3, column = 8).border = topBottom_thick_border

	return active_sheet




def createSheets(all_tables):
	wb = Workbook()
	active_sheet = wb.create_sheet("Overview",0)
	#wb.create_sheet("Abstract",1)
	active_sheet = wb.create_sheet("Summary",1)

	#Summary Sheet
	active_sheet = wb.create_sheet("Exceptions",2)
	active_sheet = styleSheet(active_sheet)

	for table in all_tables:
		table = table.replace("\\", "")
		active_sheet = wb.create_sheet(title=table[0:31].replace(":"," "))
		active_sheet.row_dimensions[3].height = 40

		active_sheet.merge_cells('A1:B1')
		active_sheet['A1'].value = "Return to Overview"
		active_sheet['A1'].hyperlink = "#'Overview'!A1"
		active_sheet['A1'].alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet['A1'].font = Font(name="CiscoSans" , size = 18, bold=True)
		active_sheet['A1'].border = thin_border


		active_sheet.merge_cells('A2:B2')
		active_sheet['A2'].value = "Return to Exception List"
		active_sheet['A2'].hyperlink = "#'Exceptions'!A1"
		active_sheet['A2'].alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet['A2'].font = Font(name="CiscoSans" , size = 18, bold=True)
		active_sheet['A2'].border = thin_border

		active_sheet.column_dimensions['A'].width = 25
		active_sheet.column_dimensions['B'].width = 25
		active_sheet.column_dimensions['C'].width = 40
		active_sheet.column_dimensions['D'].width = 25
		active_sheet.column_dimensions['E'].width = 25
		active_sheet.column_dimensions['F'].width = 80
		active_sheet.column_dimensions['G'].width = 30

		active_sheet.cell(row = 3, column = 1).value = 'NMS Area'
		active_sheet.cell(row = 3, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 1).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 1).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 2).value = 'Exception'
		active_sheet.cell(row = 3, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 2).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 2).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 3).value = 'Affected Devices'
		active_sheet.cell(row = 3, column = 3).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 3).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 3).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 4).value = 'Net Rule'
		active_sheet.cell(row = 3, column = 4).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 4).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 4).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 5).value = 'Net Advice'
		active_sheet.cell(row = 3, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 5).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 5).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 6).value = 'NCE Comments'
		active_sheet.cell(row = 3, column = 6).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 6).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 6).border = topBottom_thick_border

		active_sheet.cell(row = 3, column = 7).value = 'Additional Info'
		active_sheet.cell(row = 3, column = 7).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 3, column = 7).font = Font(name="CiscoSans" , size = 14, bold=True)
		active_sheet.cell(row = 3, column = 7).border = topBottom_thick_border


	active_sheet = wb.create_sheet(title="Critical")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="High")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Medium")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Low")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Informational")
	active_sheet = styleSheet(active_sheet)


	active_sheet = wb.create_sheet(title="Fault Management")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Configuration Management")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Capacity Management")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Performance Management")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Security Management")
	active_sheet = styleSheet(active_sheet)
	active_sheet = wb.create_sheet(title="Verify Manually")
	active_sheet = styleSheet(active_sheet)



	return wb

def insert_in_sheet(active_sheet, sheet_type, ex, sev):
	color = "FFFFFF"
	if(sev == "Critical"):
		color = "00BCEB"
	elif(sev == "High"):
		color = "E2231A"
	elif(sev == "Medium"):
		color = "FBAB18"
	elif(sev == "Low"):
		color = "EED202"
	elif(sev == "Informational"):
		color = "6ABF4B"

	if(sheet_type == 'exception'):
		insertRow = active_sheet.max_row + 1
		active_sheet.row_dimensions[insertRow].height = 100



		active_sheet.cell(row = insertRow , column = 1).value = ex["Table Name"]
		active_sheet.cell(row = insertRow , column = 1).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 1).font = Font(name="CiscoSans" , size = 11, color = "1E4471", bold=False , underline='single')
		active_sheet.cell(row = insertRow , column = 1).border = thin_border
		active_sheet.cell(row = insertRow , column = 1).hyperlink = "#'"+ex["Table Name"][0:31]+"'!A1"

		active_sheet.cell(row = insertRow , column = 2).value = ex["NMS Area"]
		active_sheet.cell(row = insertRow , column = 2).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 2).font = Font(name="CiscoSans" , size = 11, color = "1E4471", bold=False , underline='single')
		active_sheet.cell(row = insertRow , column = 2).border = thin_border
		active_sheet.cell(row = insertRow , column = 2).hyperlink = "#'"+ex["NMS Area"]+"'!A1"

		active_sheet.cell(row = insertRow , column = 3).value = ex["Exception"]
		active_sheet.cell(row = insertRow , column = 3).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 3).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 3).border = thin_border
		active_sheet.cell(row = insertRow , column = 3).fill = PatternFill(start_color=color,end_color=color,fill_type='solid')

		active_sheet.cell(row = insertRow , column = 4).value = '\n'.join(ex["Devices"])
		active_sheet.cell(row = insertRow , column = 4).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 4).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 4).border = thin_border

		active_sheet.cell(row = insertRow , column = 5).value = ex["Net Rule"]
		active_sheet.cell(row = insertRow , column = 5).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 5).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 5).border = thin_border

		active_sheet.cell(row = insertRow , column = 6).value = ex["Net Advice"]
		active_sheet.cell(row = insertRow , column = 6).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 6).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 6).border = thin_border

		active_sheet.cell(row = insertRow , column = 7).value = ex["NCE Comment"]
		active_sheet.cell(row = insertRow , column = 7).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 7).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 7).border = thin_border

		active_sheet.cell(row = insertRow , column = 8).value = '\n'.join(ex["Additional Info"])
		active_sheet.cell(row = insertRow , column = 8).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 8).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 8).border = thin_border


	elif(sheet_type == "table"):
		insertRow = active_sheet.max_row + 1
		active_sheet.row_dimensions[insertRow].height = 100


		active_sheet.cell(row = insertRow , column = 1).value = ex["NMS Area"]
		active_sheet.cell(row = insertRow , column = 1).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 1).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 1).border = thin_border


		active_sheet.cell(row = insertRow , column = 2).value = ex["Exception"]
		active_sheet.cell(row = insertRow , column = 2).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 2).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 2).border = thin_border
		active_sheet.cell(row = insertRow , column = 2).fill = PatternFill(start_color=color,end_color=color,fill_type='solid')


		active_sheet.cell(row = insertRow , column = 3).value = '\n'.join(ex["Devices"])
		active_sheet.cell(row = insertRow , column = 3).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 3).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 3).border = thin_border

		active_sheet.cell(row = insertRow , column = 4).value = ex["Net Rule"]
		active_sheet.cell(row = insertRow , column = 4).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 4).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 4).border = thin_border

		active_sheet.cell(row = insertRow , column = 5).value = ex["Net Advice"]
		active_sheet.cell(row = insertRow , column = 5).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 5).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 5).border = thin_border

		active_sheet.cell(row = insertRow , column = 6).value = ex["NCE Comment"]
		active_sheet.cell(row = insertRow , column = 6).alignment = Alignment(horizontal="center" , vertical="center", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 6).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 6).border = thin_border

		active_sheet.cell(row = insertRow , column = 7).value = '\n'.join(ex["Additional Info"])
		active_sheet.cell(row = insertRow , column = 7).alignment = Alignment(horizontal="center" , vertical="top", wrap_text=True)
		active_sheet.cell(row = insertRow , column = 7).font = Font(name="CiscoSans" , size = 11, bold=False)
		active_sheet.cell(row = insertRow , column = 7).border = thin_border


def create_summary_sheet(wb, cpyKey, auditID, audit_type, all_exceptions):

	top_exceptions = {}
	top_devices = {}
	for item in all_exceptions:
		if(item["Exception Name"] in top_exceptions.keys()):
			top_exceptions[item["Exception Name"]] = top_exceptions[item["Exception Name"]] + 1
		else:
			top_exceptions[item["Exception Name"]] = 1

		if(item["Host Name (IP Address)"] in top_devices.keys()):
			top_devices[item["Host Name (IP Address)"]] = top_devices[item["Host Name (IP Address)"]] + 1
		else:
			top_devices[item["Host Name (IP Address)"]] = 1

	top_exceptions = sorted(top_exceptions.items(), key=lambda x: x[1], reverse=True)
	top_devices = sorted(top_devices.items(), key=lambda x: x[1], reverse=True)



	active_sheet = wb['Summary']
	active_sheet.column_dimensions['A'].width = 40
	active_sheet.column_dimensions['B'].width = 40
	active_sheet.column_dimensions['C'].width = 40
	active_sheet.column_dimensions['D'].width = 40
	active_sheet.column_dimensions['E'].width = 40
	active_sheet.column_dimensions['F'].width = 40
	active_sheet.column_dimensions['G'].width = 40
	max_i = 0

	for z in range(1,50):
		active_sheet.cell(row = 1 , column = z).border = bottom_thick_border

	active_sheet.cell(row = 1, column = 1).value = 'Software Summary'
	active_sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 1, column = 1).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = 1, column = 1).border = thick_border

	active_sheet.cell(row = 1, column = 5).value = 'Hardware Summary'
	active_sheet.cell(row = 1, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = 1, column = 5).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = 1, column = 5).border = thick_border

	try:
		software_table = list(jsonCreator.getFromDB(cpyKey, auditID, 'softwareTable'))
	except Exception as e:
		print(str(e))
	if(software_table):
		hardware_details = {}
		software_details = {}
		for jsonData in software_table:
			del jsonData["_id"]
			jsonData["Host Name (IP Address)"] = jsonData["Host Name (IP Address)"].replace("\xa0","")
			if(jsonData["Software Version"] in software_details.keys()):
				software_details[jsonData["Software Version"]] = software_details[jsonData["Software Version"]] + 1
			else:
				software_details[jsonData["Software Version"]] = 1
			if(jsonData["Product Type"] in hardware_details.keys()):
				hardware_details[jsonData["Product Type"]] = hardware_details[jsonData["Product Type"]] + 1
			else:
				hardware_details[jsonData["Product Type"]] = 1
		max_i = 0
		active_sheet.cell(row=2, column=1).value = "software version"
		active_sheet.cell(row = 2, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 2, column = 1).font = Font(name="CiscoSans" , size = 12, bold=True)


		active_sheet.cell(row=2, column=2).value = "count"
		active_sheet.cell(row = 2, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 2, column = 2).font = Font(name="CiscoSans" , size = 12, bold=True)


		i = 3
		for key,value in software_details.items():
			active_sheet.cell(row=i, column=1).value = key
			active_sheet.cell(row=i, column=2).value = value
			i = i + 1
		max_i = i
		active_sheet.cell(row=2, column=5).value = "Product type"
		active_sheet.cell(row = 2, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 2, column = 5).font = Font(name="CiscoSans" , size = 12, bold=True)


		active_sheet.cell(row=2, column=6).value = "count"
		active_sheet.cell(row = 2, column = 6).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 2, column = 6).font = Font(name="CiscoSans" , size = 12, bold=True)

		
		i = 3
		for key,value in hardware_details.items():
			active_sheet.cell(row=i, column=5).value = key
			active_sheet.cell(row=i, column=6).value = value
			i = i + 1
		if(i > max_i):
				max_i = i
	
		if max_i < 20:
			max_i = 20

		### PIE Chart for Hardware Summary
		pie = PieChart()
		labels = Reference(active_sheet, min_col=5, min_row=3, max_row=len(hardware_details)+3)
		data = Reference(active_sheet, min_col=6, min_row=2, max_row=len(hardware_details)+2)
		pie.add_data(data, titles_from_data=True)
		pie.set_categories(labels)
		pie.title = "Hardware Summary"
		active_sheet.add_chart(pie,"E3")
		
		### Column Chart for Software Summary

		chart1 = BarChart()
		chart1.type = "col"
		chart1.style = 10
		chart1.title = "Software Summary"
		chart1.y_axis.title = 'Version'
		chart1.x_axis.title = 'Count'
	
		data = Reference(active_sheet, min_col=2, min_row=2, max_row=len(software_details)+2)
		cats = Reference(active_sheet, min_col=1, min_row=3, max_row=len(software_details)+3)
		chart1.add_data(data, titles_from_data=True)
		chart1.set_categories(cats)
		chart1.shape = 2
		active_sheet.add_chart(chart1, "A3")


	if(software_table):
		k= max_i + 1
	else:
		k = 20
		active_sheet.merge_cells('B10:E10')
		active_sheet.cell(row = 10, column = 2).value = "Software and Hardware summary is not available as NP group was not selected during ZIP Upload on NAA Portal"
		active_sheet.cell(row = 10, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
		active_sheet.cell(row = 10, column = 2).font = Font(name="CiscoSans" , size = 14, bold=True, color = "000080")

	for z in range(1,50):
		active_sheet.cell(row = k , column = z).border = bottom_thick_border

	active_sheet.cell(row = k, column = 1).value = 'Severity Breakdown'
	active_sheet.cell(row = k, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 1).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = k, column = 1).border = thick_border

	active_sheet.cell(row = k, column = 5).value = 'FCCAPS Breakdown'
	active_sheet.cell(row = k, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 5).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = k, column = 5).border = thick_border

	k = k + 1

	active_sheet.cell(row=k, column=1).value = "Severity"
	active_sheet.cell(row = k, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 1).font = Font(name="CiscoSans" , size = 12, bold=True)


	active_sheet.cell(row=k, column=2).value = "Count"
	active_sheet.cell(row = k, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 2).font = Font(name="CiscoSans" , size = 12, bold=True)


	crit = wb["Overview"].cell(row = wb["Overview"].max_row , column = 2).value
	crit = crit.split("(")
	
	high = wb["Overview"].cell(row = wb["Overview"].max_row , column = 3).value
	high = high.split("(")

	med = wb["Overview"].cell(row = wb["Overview"].max_row , column = 4).value
	med = med.split("(")

	lw = wb["Overview"].cell(row = wb["Overview"].max_row , column = 5).value
	lw = lw.split("(")

	inf = wb["Overview"].cell(row = wb["Overview"].max_row , column = 6).value
	inf =inf.split("(")
	
	active_sheet.cell(row=k+1, column=1).value = "Critical"
	active_sheet.cell(row=k+1, column=2).value = crit[0]+"(Overview!"+crit[1]
	active_sheet.cell(row=k+2, column=1).value = "High"
	active_sheet.cell(row=k+2, column=2).value = high[0]+"(Overview!"+high[1]
	active_sheet.cell(row=k+3, column=1).value = "Medium"
	active_sheet.cell(row=k+3, column=2).value = med[0]+"(Overview!"+med[1]	
	active_sheet.cell(row=k+4, column=1).value = "Low"
	active_sheet.cell(row=k+4, column=2).value = lw[0]+"(Overview!"+lw[1]	
	active_sheet.cell(row=k+5, column=1).value = "Informational"
	active_sheet.cell(row=k+5, column=2).value = inf[0]+"(Overview!"+inf[1]

	active_sheet.cell(row=k, column=5).value = "NMS Area"
	active_sheet.cell(row = k, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 5).font = Font(name="CiscoSans" , size = 12, bold=True)

	active_sheet.cell(row=k, column=6).value = "Count"
	active_sheet.cell(row = k, column = 6).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 6).font = Font(name="CiscoSans" , size = 12, bold=True)

	active_sheet.cell(row=k+1, column=5).value = "Fault Management"
	active_sheet.cell(row=k+1, column=6).value = wb["Overview"].cell(row = 9 , column = 2).value
	active_sheet.cell(row=k+2, column=5).value = "Capacity Management"
	active_sheet.cell(row=k+2, column=6).value = wb["Overview"].cell(row = 9 , column = 4).value
	active_sheet.cell(row=k+3, column=5).value = "Configuration Management"
	active_sheet.cell(row=k+3, column=6).value = wb["Overview"].cell(row = 9 , column = 3).value	
	active_sheet.cell(row=k+4, column=5).value = "Performance Management"
	active_sheet.cell(row=k+4, column=6).value = wb["Overview"].cell(row = 9 , column = 5).value
	active_sheet.cell(row=k+5, column=5).value = "Security Management"
	active_sheet.cell(row=k+5, column=6).value = wb["Overview"].cell(row = 9 , column = 6).value

	### PIE Chart for Severity Breakdown
	pie1 = PieChart()
	labels1 = Reference(active_sheet, min_col=1, min_row=k+1, max_row=k+6)
	data1 = Reference(active_sheet, min_col=2, min_row=k, max_row=k+5)
	pie1.add_data(data1, titles_from_data=True)
	pie1.set_categories(labels1)
	pie1.title = "Severity Breakdown"
	active_sheet.add_chart(pie1,"A"+str(k+1))

	### Column Chart for FCCAPS Breakdown
	chart4 = BarChart()
	chart4.type = "col"
	chart4.style = 10
	chart4.title = "FCCAPS Breakdown"
	chart4.y_axis.title = 'Count'
	chart4.x_axis.title = 'NMS Area'
	
	data3 = Reference(active_sheet, min_col=6, min_row=k, max_row=k+5)
	cats3 = Reference(active_sheet, min_col=5, min_row=k+1, max_row=k+6)
	chart4.add_data(data3, titles_from_data=True)
	chart4.set_categories(cats3)
	chart4.shape = 2
	active_sheet.add_chart(chart4, "E"+str(k+1))


	k = (k * 2) + 1

	for z in range(1,50):
		active_sheet.cell(row = k , column = z).border = bottom_thick_border

	active_sheet.cell(row = k, column = 1).value = 'Top 10 Devices'
	active_sheet.cell(row = k, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 1).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = k, column = 1).border = thick_border

	active_sheet.cell(row = k, column = 5).value = 'Top 10 Exceptions'
	active_sheet.cell(row = k, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 5).font = Font(name="CiscoSans" , size = 16, bold=True, color = "000080")
	active_sheet.cell(row = k, column = 5).border = thick_border
	

	k = k + 1

	active_sheet.cell(row=k, column=1).value = "Device Name"
	active_sheet.cell(row = k, column = 1).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 1).font = Font(name="CiscoSans" , size = 12, bold=True)

	active_sheet.cell(row=k, column=2).value = "Count"
	active_sheet.cell(row = k, column = 2).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 2).font = Font(name="CiscoSans" , size = 12, bold=True)


	active_sheet.cell(row=k, column=5).value = "Exception Name"
	active_sheet.cell(row = k, column = 5).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 5).font = Font(name="CiscoSans" , size = 12, bold=True)

	active_sheet.cell(row=k, column=6).value = "Count"
	active_sheet.cell(row = k, column = 6).alignment = Alignment(horizontal="center" , vertical="center")
	active_sheet.cell(row = k, column = 6).font = Font(name="CiscoSans" , size = 12, bold=True)


	top_dev_range = 10
	top_exp_range = 10

	if(len(top_devices) < 10):
		top_dev_range = len(top_devices)
	if(len(top_exceptions) < 10):
		top_exceptions = len(top_exceptions)

	for i in range(0,top_dev_range):
		active_sheet.cell(row = k+i+1 , column = 1).value = top_devices[i][0]
		active_sheet.cell(row = k+i+1 , column = 2).value = top_devices[i][1]

	for i in range(0,top_exp_range):
		active_sheet.cell(row = k+i+1 , column = 5).value = top_exceptions[i][0]
		active_sheet.cell(row = k+i+1 , column = 6).value = top_exceptions[i][1]


	### Column Chart for Top 10 Devices
	chart5 = BarChart()
	chart5.type = "col"
	chart5.style = 10
	chart5.title = "Top 10 Devices (with highest exceptions)"
	chart5.y_axis.title = 'Count'
	chart5.x_axis.title = 'Device Name'
	
	data_top_dev = Reference(active_sheet, min_col=2, min_row=k, max_row=k+top_dev_range)
	ref_top_dev = Reference(active_sheet, min_col=1, min_row=k+1, max_row=k+top_dev_range+1)
	chart5.add_data(data_top_dev, titles_from_data=True)
	chart5.set_categories(ref_top_dev)
	chart5.shape = 2
	active_sheet.add_chart(chart5, "A"+str(k+1))


	### Column Chart for Top 10 Exceptions
	chart6 = BarChart()
	chart6.type = "col"
	chart6.style = 10
	chart6.title = "Top 10 Exceptions (Occuring maximum times)"
	chart6.y_axis.title = 'Count'
	chart6.x_axis.title = 'Exception Name'
	
	data_top_exp = Reference(active_sheet, min_col=6, min_row=k, max_row=k+top_exp_range)
	ref_top_exp = Reference(active_sheet, min_col=5, min_row=k+1, max_row=k+top_exp_range+1)
	chart6.add_data(data_top_exp, titles_from_data=True)
	chart6.set_categories(ref_top_exp)
	chart6.shape = 2
	active_sheet.add_chart(chart6, "E"+str(k+1))


def createReport():
	try:
		cpyKey , auditID , audit_type, Customer_Name, email_id = getAuditDetails()
		all_exceptions = list(jsonCreator.getFromDB(cpyKey, auditID, 'allExceptions'))
		severity_breakdown = list(jsonCreator.getFromDB(cpyKey, auditID, 'severityBreakdown'))
		all_ex = []

		#only take exceptions which need to be reported
		for exception in all_exceptions:
			if(exception["DNR_Flag"] == "0"):
				all_ex.append(exception)
		all_exceptions = all_ex


		all_data = list(jsonCreator.getFromDB(cpyKey, auditID, 'allData'))
		
		#print(severity_breakdown)
		audit_information = list(jsonCreator.getFromDB(cpyKey, auditID, 'audit_information'))
		if not audit_information:
			audit_information = [{"Customer_Name": Customer_Name, "Collection Start Time": "NA", "Collection End Time": "NA", "Collection Period (Days)": "NA", "Devices Attempted": "NA", "Devices Passed": "NA", "Devices Failed/Excluded": "NA", "Devices Duplicated": "NA"}]
		exception_list = {"Critical" : {} , "High" : {} , "Medium" : {} , "Low" : {} , "Informational" : {} }
	
	
		all_tables = {}
		for exception in all_exceptions:
	
			if(exception["Table Name"] in all_tables.keys()):
				if(exception["Severity"] in all_tables[exception["Table Name"]].keys()):
					all_tables[exception["Table Name"]][exception["Severity"]] = all_tables[exception["Table Name"]][exception["Severity"]] + 1
				else:
					all_tables[exception["Table Name"]][exception["Severity"]] = 1
			else:
				all_tables[exception["Table Name"]] = {}
				all_tables[exception["Table Name"]][exception["Severity"]] = 1
	
	
	
			if(exception["Exception Name"] in exception_list[exception["Severity"]].keys()):
				if(exception["Host Name (IP Address)"].split(" ")[0].strip() not in exception_list[exception["Severity"]][exception["Exception Name"]]["Devices"]):
					exception_list[exception["Severity"]][exception["Exception Name"]]["Devices"].append(exception["Host Name (IP Address)"].split(" ")[0].strip())
					exception_list[exception["Severity"]][exception["Exception Name"]]["Additional Info"].append(exception["Host Name (IP Address)"].split(" ")[0]+" - "+exception["Exception Value"])
	
			else:
				netRule = "No Net Rule available"
				netAdvice = "No Net Advice available"
				NetRule_NetAdvice = getNetRulesv2(cpyKey, auditID, audit_type, exception["Table Name"], exception["Exception Name"])
				#NetRule_NetAdvice = getNetRules(audit_type, exception["Table Name"], exception["Exception Name"])
				if(len(NetRule_NetAdvice) > 0):
					netRule = NetRule_NetAdvice[0]["net_rule"]
					netAdvice = NetRule_NetAdvice[0]["net_advice"]
	
				tempJson = {"Table Name" : exception["Table Name"] , "NMS Area" : exception["NMS Area"] , "Exception" : exception["Exception Name"] , "Devices" : [exception["Host Name (IP Address)"].split(" ")[0].strip()] , "Net Rule" : netRule ,"Net Advice" : netAdvice ,"NCE Comment" : exception["NCE_Comment"] , "Additional Info" : [exception["Host Name (IP Address)"].split(" ")[0]+" - "+exception["Exception Value"]]}
				exception_list[exception["Severity"]][exception["Exception Name"]] = tempJson

		wb = createSheets(all_tables)
		for exception in exception_list.keys():
			for key in exception_list[exception].keys():
				active_sheet = wb["Exceptions"]
				insert_in_sheet(active_sheet,"exception", exception_list[exception][key], exception)
				active_sheet = wb[exception_list[exception][key]["Table Name"][0:31].replace(":"," ")]
				insert_in_sheet(active_sheet, "table", exception_list[exception][key] , exception)
				active_sheet = wb[exception]
				insert_in_sheet(active_sheet,"exception", exception_list[exception][key] , exception)
				active_sheet = wb[exception_list[exception][key]["NMS Area"]]
				insert_in_sheet(active_sheet,"exception", exception_list[exception][key] , exception)
	
	
		if not os.path.exists(scriptPath+"/Reports"):
			os.makedirs(scriptPath+"/Reports")
	
	
		create_overview_sheet(wb, audit_information, all_tables, severity_breakdown)
		create_summary_sheet(wb,cpyKey , auditID, audit_type, all_exceptions)
	
		all_sheets = wb.sheetnames
	
		for entry in all_data:
			
			if(entry["Table Name"][0:31] in all_sheets and entry["Table Name"] in all_tables.keys()):
				active_sheet = wb[entry["Table Name"][0:31]]
				if("cols" in all_tables[entry["Table Name"]].keys()):
					if(entry["Column Name"] in all_tables[entry["Table Name"]]["cols"]):
						column_number = all_tables[entry["Table Name"]]["cols"].index(entry["Column Name"])
					else:
						all_tables[entry["Table Name"]]["cols"].append(entry["Column Name"])
						column_number = all_tables[entry["Table Name"]]["cols"].index(entry["Column Name"])
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).value = entry["Column Name"]
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).alignment = Alignment(horizontal="center" , vertical="center")
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).font = Font(name="CiscoSans" , size = 14, bold=True)
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).border = topBottom_thick_border
	
				else:
					if("max_row" in all_tables[entry["Table Name"]].keys()):
						continue
					else:
						all_tables[entry["Table Name"]]["max_row"] = active_sheet.max_row
						all_tables[entry["Table Name"]]["cols"] =["Host Name"]
						for k in range(1,50):
							active_sheet.cell(row = all_tables[entry["Table Name"]]["max_row"]+2, column = k).border = bottom_thick_border
	
						active_sheet.row_dimensions[all_tables[entry["Table Name"]]["max_row"]+5].height = 50
						
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=1).value = "Host Name"
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=1).alignment = Alignment(horizontal="center" , vertical="center")
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=1).font = Font(name="CiscoSans" , size = 14, bold=True)
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=1).border = topBottom_thick_border
	
						all_tables[entry["Table Name"]]["cols"].append(entry["Column Name"])
						column_number = all_tables[entry["Table Name"]]["cols"].index(entry["Column Name"])
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).value = entry["Column Name"]
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).alignment = Alignment(horizontal="center" , vertical="center")
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).font = Font(name="CiscoSans" , size = 14, bold=True)
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5, column=column_number+1).border = topBottom_thick_border
	
	
				active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=1).value = entry["Host Name (IP Address)"]
				active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).value = entry["Column Value"]
				sev = ""
				for data in all_exceptions:
					if(entry["Table Name"] == data["Table Name"] and entry["Row Number"] == data["Row Number"] and entry["Column Name"] == data["Exception Name"] and entry["Column Value"] == data["Exception Value"]):
						sev = data["Severity"]
	
				if(sev):
					if(sev == "Critical"):
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).fill = PatternFill(start_color='00BCEB',end_color='00BCEB',fill_type='solid')
					elif(sev=="High"):
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).fill = PatternFill(start_color='E2231A',end_color='E2231A',fill_type='solid')
					elif(sev=="Medium"):
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).fill = PatternFill(start_color='FBAB18',end_color='FBAB18',fill_type='solid')
					elif(sev=="Low"):
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).fill = PatternFill(start_color='EED202',end_color='EED202',fill_type='solid')
					elif(sev=="Informational"):
						active_sheet.cell(row=all_tables[entry["Table Name"]]["max_row"]+5+entry["Row Number"],column=column_number+1).fill = PatternFill(start_color='6ABF4B',end_color='6ABF4B',fill_type='solid')
	
				
		#print(all_tables)
	
	
		wb.active = 0
		del wb["Sheet"]
		filename = cpyKey+"_"+auditID
		wb.save(scriptPath+"/Reports/"+filename+".xlsx")
		wb.close()

		body = """Hi, 
					Thank you for using Network Audit Analyser tool. The excel sheet report is attached with this mail.
					Please reach out to bgl-cx-bcs-audit@cisco.com if you have any questions
					Thank you,
					BCS Audit team """
		
		body_bot = """Thank you for using Network Audit Analyser tool. The excel sheet report for audit '"""+str(auditID)+"""' is attached. Please reach out to bgl-cx-bcs-audit@cisco.com if you have any questions"""
		
		mail_utility.send_email(email_id, 'Please do not reply to this mail.',body,scriptPath+"/Reports/"+filename+".xlsx")

		try:
			api = WebexTeamsAPI(access_token='YzY4ZjFjY2EtOWY4Ny00NmRmLWE5MTUtZWFlYzJiM2MwMWJlYTgyMzVlODEtNWI0_PF84_1eb65fdf-9643-417f-9974-ad72cae0e10f')
			api.messages.create(toPersonEmail=str(sys.argv[5]) , text = body_bot , files = [scriptPath+"/Reports/"+filename+".xlsx"])

		except Exception as e:
			print("Error sending file via Webex Bot.\n")
			print(str(e))
	

	except Exception as e:
		print("Exception Occured ::: " + str(e))

		body = """Hi, 
					There is some error generating report. Please contact bgl-cx-bcs-audit@cisco.com to resolve the issue.

					Regards,
					BCS Audit team """
		
		body_bot = """Hi, There is some error generating report for audit id '"""+str(auditID)+"""'. Please contact bgl-cx-bcs-audit@cisco.com to resolve the issue.
		Regards,
		BCS Audit team """

		mail_utility.send_email(email_id, 'Please do not reply to this mail.',body,'')
		try:
			api = WebexTeamsAPI(access_token='YzY4ZjFjY2EtOWY4Ny00NmRmLWE5MTUtZWFlYzJiM2MwMWJlYTgyMzVlODEtNWI0_PF84_1eb65fdf-9643-417f-9974-ad72cae0e10f')
			api.messages.create(toPersonEmail=str(sys.argv[5]) , text = body_bot)

		except Exception as e:
			print("Error sending file via Webex Bot.\n")
			print(str(e))



body1 = """Hi, 
			Excel report generation initiated for audit ID """+str(sys.argv[2])+""". Please allow upto 12 hours before raising 
			issue with bgl-cx-bcs-audit@cisco.com

			Regards,
			BCS Audit team """
mail_utility.send_email(str(sys.argv[5]), 'Please do not reply to this mail.',body1,'')
createReport()
