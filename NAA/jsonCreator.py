#!/usr/bin/env python3
"""
Author: Pushkaraj Lahankar <plahanka@cisco.com>

Color coding for flagging exceptions:
Critical : #ADDFFF -> Blue
High : #FF0000 -> Red
Medium : #F9966B -> Orange
Low : #FFFF00 -> Yellow

"""
import os
import sys
import datetime
from zipfile import ZipFile
import shutil
import glob
from bs4 import BeautifulSoup
import json
import pymongo
import re
import openpyxl
import cna_graph_images
import mail_utility


os.environ["PYTHONIOENCODING"] = "utf-32"

cssStatement = '<link rel="stylesheet" type="text/css" href="reportStylesheet.css">'
scriptPath = os.path.dirname(os.path.realpath(__file__))
headlesstableFlag = 0  # This is being used to pull data for audit_information. This table does not have any headers. But there can be multiple such tables in


# audti_summary. This is the very first table in audit_summary. Making Flag = 1 will ensure this is not overwritten by other headless tables.

def getTime(format):
    if (format == "timestamp"):
        return (str(datetime.datetime.timestamp(datetime.datetime.now())).split(".")[0])
    elif (format == "date"):
        return (datetime.datetime.now())


def getBasicInfoFromDB(cpy_key, audit_id):
    jsonData = {}
    result = []
    dbName = "CNA_Visualizer"
    dbClient = pymongo.MongoClient('localhost', 27017)
    db = dbClient[dbName]
    collection = db['upload_information']
    if (audit_id == "blank"):
        query = {"customer_key": cpy_key}
        result = db['upload_information'].find(query)
        jsonData['customer_name'] = result[0]['customer_name']
    else:
        query = {"customer_key": cpy_key, "audit_id": audit_id}
        result = db['upload_information'].find(query)
        jsonData['customer_name'] = result[0]['customer_name']
        jsonData['audit_type'] = result[0]['audit_type']
    for res in result:
        try:
            jsonData['audit_ids'].append(res['audit_id'])
        except KeyError:
            jsonData['audit_ids'] = [res['audit_id']]

    jsonData["audit_ids"] = list(set(jsonData["audit_ids"]))
    print(jsonData)
    dbClient.close()
    return jsonData


def getFromDB(cpyKey, auditID, jsonFor):
    dbName = "CNA_Visualizer"
    dbClient = pymongo.MongoClient('localhost', 27017)
    db = dbClient[dbName]
    collection = db[cpyKey]
    query = {"jsonFor": jsonFor, "Audit_ID": auditID}
    result = collection.find(query)
    dbClient.close()
    return result


def insertBasicInfo(jsonData):
    dbName = "CNA_Visualizer"
    dbClient = pymongo.MongoClient('localhost', 27017)
    db = dbClient[dbName]
    collection = db['upload_information']
    ins = collection.insert_one(jsonData)
    del jsonData["_id"]
    dbClient.close()


def insertIntoDB(cpyKey, auditID, customerName, audit_type):
    dbName = "CNA_Visualizer"
    dbClient = pymongo.MongoClient('localhost', 27017)
    db = dbClient[dbName]
    collection = db[cpyKey]

    customerName = customerName.replace(" ", "_")
    for f in os.listdir(scriptPath + "/ExceptionTables" + auditID + "/data"):
        if (f.endswith(".json")):
            jsonData = []
            with open(scriptPath + "/ExceptionTables" + auditID + "/data/" + f) as fileptr:
                jsonData = json.loads(fileptr.read())

            # Add metadata to every JSON Object.

            if (isinstance(jsonData, dict) and bool(jsonData)):
                print(f)
                print("In insertInDB :::  " + str(jsonData))
                jsonData["jsonFor"] = f.split(".")[0]
                jsonData["Customer_Name"] = customerName
                jsonData["Audit_ID"] = auditID
                jsonData["Audit_Type"] = audit_type
                jsonData["timeStamp"] = getTime("timestamp")
                if ("allExceptions" in f):
                    jsonData["NCE_Comment"] = ""
                    jsonData["History"] = []
                    jsonData["DNR_Flag"] = "0"
                ins = collection.insert_one(jsonData)
                del jsonData["_id"]
            elif (isinstance(jsonData, list) and jsonData):
                for data in jsonData:
                    data["jsonFor"] = f.split(".")[0]
                    data["Customer_Name"] = customerName
                    data["Audit_ID"] = auditID
                    data["Audit_Type"] = audit_type
                    data["timeStamp"] = getTime("timestamp")
                    if ("allExceptions" in f):
                        data["NCE_Comment"] = ""
                        data["History"] = []
                        data["DNR_Flag"] = "0"

                ins = collection.insert_many(jsonData)
                for d in jsonData:
                    # ins = collection.insert_one(d)
                    del d["_id"]
    dbClient.close()


def createIOSXRSoftwareHardwareGraph(audit_id):
    softwareSummary = {}
    swSummary = []
    hwSummary = []
    hardwareSummary = {}
    softwareTable = []
    jsonData = []
    try:
        with open(scriptPath + "/ExceptionTables" + audit_id + "/json/IOS XR Software Summary Table.json", "r") as f:
            jsonData = json.loads(f.read())
        data = jsonData["data"]
        exception = {}
        for item in data:
            if (list(item.keys())[0] == "Version"):
                exception["Software_Version"] = item["Version"]["value"]
            elif (list(item.keys())[0] == "Total"):
                exception["Count"] = item["Total"]["value"]
            elif (list(item.keys())[0] == "Version %"):
                swSummary.append(exception)
                exception = {}
    except Exception as e:
        print(str(e))

    try:
        with open(scriptPath + "/ExceptionTables" + audit_id + "/json/Hardware Summary Table.json", "r") as f:
            jsonData = json.loads(f.read())
        data = jsonData["data"]
        exception = {}
        for item in data:
            if (list(item.keys())[0] == "Model"):
                exception["Chassis_Type"] = item["Model"]["value"]
            elif (list(item.keys())[0] == "Total"):
                exception["Count"] = item["Total"]["value"]
            elif (list(item.keys())[0] == "Model %"):
                hwSummary.append(exception)
                exception = {}
    except Exception as e:
        print(str(e))

    try:
        with open(scriptPath + "/ExceptionTables" + audit_id + "/json/Router Software Table.json", "r") as f:
            jsonData = json.loads(f.read())
        data = jsonData["data"]
        exception = {}
        for item in data:
            if (list(item.keys())[0] == "Host Name (IP Address)"):
                exception["Host Name (IP Address)"] = item["Host Name (IP Address)"]["value"]
            elif (list(item.keys())[0] == "Model"):
                exception["Chassis_Type"] = item["Model"]["value"]
            elif (list(item.keys())[0] == "IOS XR Version"):
                exception["Software_Version"] = item["IOS XR Version"]["value"]
            elif (list(item.keys())[0] == "NREPs"):
                softwareTable.append(exception)
                exception = {}
    except Exception as e:
        print(str(e))

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/softwareTable.json", "w") as f:
        f.write(json.dumps(softwareTable))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/softwareSummaryGraph.json", "w") as f:
        f.write(json.dumps(swSummary))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/hardwareSummaryGraph.json", "w") as f:
        f.write(json.dumps(hwSummary))


def createCPUandMemoryTable(audit_id):
    CPU_Memory_Table = []
    CPU_Sorted = []
    Memory_Sorted = []
    jsonData = []
    try:
        with open(scriptPath + "/ExceptionTables" + audit_id + "/json/CPU Table.json", "r") as f:
            jsonData = json.loads(f.read())
    except Exception as e:
        print(str(e))

    data = jsonData["data"]
    exception = {}
    for item in data:
        if (list(item.keys())[0] == "Host Name (IP Address)" and len(data) > 1):
            exception["Host Name (IP Address)"] = item[list(item.keys())[0]]["value"].replace("\xa0", " ")
        elif (list(item.keys())[0] == "NREPs"):
            CPU_Memory_Table.append(exception)
            exception = {}
        else:
            exception[list(item.keys())[0]] = item[list(item.keys())[0]]["value"]

    try:
        with open(scriptPath + "/ExceptionTables" + audit_id + "/json/Memory Table.json", "r") as f:
            jsonData = json.loads(f.read())
    except Exception as e:
        print(str(e))

    data = jsonData["data"]

    for item in data:
        if (list(item.keys())[0] != "Host Name (IP Address)" and list(item.keys())[0] != "NREPs" and len(data) > 1):

            for entry in CPU_Memory_Table:
                if (entry["Host Name (IP Address)"].replace("   ", " ") == item[list(item.keys())[0]]["device_name"] or
                        entry["Host Name (IP Address)"] == item[list(item.keys())[0]]["device_name"]):
                    entry[list(item.keys())[0]] = item[list(item.keys())[0]]["value"]

    # CPU_Sorted = sorted(CPU_Memory_Table, key = lambda i: i["Average CPU %"],reverse=True)
    # Memory_Sorted = sorted(CPU_Memory_Table, key = lambda i: i["Free Memory(%)"],reverse=False)

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/CPU_Memory_Table.json", "w") as f:
        f.write(json.dumps(CPU_Memory_Table))

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/CPU_Sorted_Table.json", "w") as f:
        f.write(json.dumps(CPU_Memory_Table))

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/Memory_Sorted_Table.json", "w") as f:
        f.write(json.dumps(CPU_Memory_Table))


def createSoftwareTable_Graphs(audit_id):
    softwareSummary = {}
    swSummary = []
    hwSummary = []
    hardwareSummary = {}
    softwareTable = []
    jsonData = []
    with open(scriptPath + "/ExceptionTables" + audit_id + "/json/Software Table.json", "r") as f:
        jsonData = json.loads(f.read())
    data = jsonData["data"]
    for item in data:
        if (list(item.keys())[0] == "SW Version" and len(data) > 1):
            if (item[list(item.keys())[0]]["value"] in list(softwareSummary.keys())):
                softwareSummary[item[list(item.keys())[0]]["value"]] = softwareSummary[
                                                                           item[list(item.keys())[0]]["value"]] + 1
            else:
                softwareSummary[item[list(item.keys())[0]]["value"]] = 1

        if (list(item.keys())[0] == "Chassis Type" and len(data) > 1):
            if (item[list(item.keys())[0]]["value"] in list(hardwareSummary.keys())):
                hardwareSummary[item[list(item.keys())[0]]["value"]] = hardwareSummary[
                                                                           item[list(item.keys())[0]]["value"]] + 1
            else:
                hardwareSummary[item[list(item.keys())[0]]["value"]] = 1

    for key in softwareSummary.keys():
        # print(key)
        tempJson = {"Software_Version": key, "Count": softwareSummary[key]}
        swSummary.append(tempJson)
    for key in hardwareSummary.keys():
        tempJson = {"Chassis_Type": key, "Count": hardwareSummary[key]}
        hwSummary.append(tempJson)

    exception = {}
    for item in data:
        if (list(item.keys())[0] == "Host Name (IP Address)"):
            exception[list(item.keys())[0]] = item[list(item.keys())[0]]["value"]
        elif (list(item.keys())[0] == "Chassis Type"):
            exception["Product Type"] = item["Chassis Type"]["value"]
        elif (list(item.keys())[0] == "SW Version"):
            exception["Software Version"] = item["SW Version"]["value"]
        elif (list(item.keys())[0] == "NREPS"):
            softwareTable.append(exception)
            exception = {}

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/softwareTable.json", "w") as f:
        f.write(json.dumps(softwareTable))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/softwareSummaryGraph.json", "w") as f:
        f.write(json.dumps(swSummary))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/hardwareSummaryGraph.json", "w") as f:
        f.write(json.dumps(hwSummary))


def getTime(format):
    if (format == "timestamp"):
        return (str(datetime.datetime.timestamp(datetime.datetime.now())).split(".")[0])
    elif (format == "date"):
        return (datetime.datetime.now())


# def getZipFile(dir):
#   print("Getting ZIP file")
#   for file in os.listdir(dir):
#       if(file.endswith(".zip")):
#           return file

def extractZip(fileName):
    with ZipFile(fileName, 'r') as zip:
        print('Extracting all the files now...')
        zip.extractall()


def getChassisName(device_name, audit_id):
    chassis_name = ""
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/deviceDetails.json", "r") as f:
        jsonData = json.loads(f.read())
        for data in jsonData:
            if (data["node name"] == device_name):
                chassis_name = data["Model"]

    return chassis_name


def device_to_chassis_mapping(scriptPath, fileName):
    nodeNames = []
    os.chdir(scriptPath + "/" + fileName + "/Full")
    for d in os.listdir("."):
        if os.path.isdir(d) and d != "summary" and d != "commonImages":
            for file in os.listdir(d):
                if (file.endswith(".html")):
                    with open(d + "/" + file, "r") as f:
                        fileData = str(f.readlines())
                        startInd = fileData.find("Node Name:")
                        endInd = fileData.find("</b>", startInd)
                        nodeName = (fileData[startInd:endInd]).split(":")[1]
                        # print(nodeName)
                        startInd2 = fileData.find("Model:")
                        endInd2 = fileData.find("</b>", startInd2)
                        Model = (fileData[startInd2:endInd2]).split(":")[1]
                        # print(Model)
                        nodeNames.append({"node name": nodeName, "Model": Model})
    return (nodeNames)


def extractTable(file_name):
    data = ""
    with open(file_name, "r") as f:
        data = f.read()
    if (len(data) > 1):
        return (data[data.find('<TABLE'): data.find('</TABLE>') + 8])
    else:
        return data


def getNMSArea(tableName, audit_id):
    NMSArea = ""
    tableName = tableName.strip()
    print("in getNMSArea :: " + tableName)
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/tableToNMSMap.json", "r") as f:
        jsonData = json.loads(f.read())
    for key in jsonData.keys():
        tableList = jsonData[key]
        if (tableName in tableList):
            NMSArea = key
            print("table name " + tableName + " found in NMS Area " + key)
            break
        elif (NMSArea == ""):
            for item in tableList:
                if ("VSAN table" in tableName):
                    if (tableName.replace("VSAN table", "Show VSAN table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("IPv4 - Router Switching Path Status (Non-Distributed Platform)" in tableName):
                    if (tableName.replace("IPv4 - Router Switching Path Status (Non-Distributed Platform)",
                                          "IPv4 - Router Switching Path Status (Non-Distributed Switching Platform)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("NVE Interface Status Table" in tableName):
                    if (tableName.replace("NVE Interface Status Table", "Interface nve status table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("IPv6 - Router Switching Path Status (Non-Distributed Platform)" in tableName):
                    if (tableName.replace("IPv6 - Router Switching Path Status (Non-Distributed Platform)",
                                          "IPv6 - Router Switching Path Status (Non-Distributed Switching Platform)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("IPv4 - Not-CEF-Switched Statistics (Non-distributed Switching Platform)" in tableName):
                    NMSArea = "Capacity Management"
                    print(NMSArea)
                    break

                elif ("IPv4 - Hot Standby Routing Protocol Table" in tableName):
                    NMSArea = "Configuration Management"
                    print(NMSArea)
                    break
                elif ("Summarization Related (ABRs only) - Main Domain" in tableName):
                    if (tableName.replace("Summarization Related (ABRs only) - Main Domain",
                                          "Summarization Related - Main Domain (ABRs only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("Summarization Related (ABRs only) - Other Domains" in tableName):
                    if (tableName.replace("Summarization Related (ABRs only) - Other Domains",
                                          "Summarization Related - Other Domain (ABRs only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("ADMIN Distance Table (Exceptions only) - Main Domain" in tableName):
                    if (tableName.replace("ADMIN Distance Table (Exceptions only) - Main Domain",
                                          "ADMIN Distance Table - Main Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("ADMIN Distance Table (Exceptions only) - Other Domains" in tableName):
                    if (tableName.replace("ADMIN Distance Table (Exceptions only) - Other Domains",
                                          "ADMIN Distance Table - Other Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("Catalyst 6500 Series MLS Netflow Configuration Table" in tableName):
                    if (tableName.replace("Catalyst 6500 Series MLS Netflow Configuration Table",
                                          "Catalyst 6500 and Cisco 7600 Series MLS Netflow Configuration Table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("Suspects Port Table" in tableName):
                    if (tableName.replace("Suspects Port Table", "Suspect Ports Table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("SIP SPA Compatibility Table" in tableName):
                    if (tableName.replace("SIP SPA Compatibility Table", "SIP and SPA Information Table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break


                elif ("Cisco 6500 7600 MPLS Interface Table" in tableName):
                    if (tableName.replace("Cisco 6500 7600 MPLS Interface Table",
                                          "Cisco 6500/7600 MPLS Interface Table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break


                elif ("Cisco 7600 MPLS Routes VRF's HW mode Table" in tableName):
                    if (tableName.replace("Cisco 7600 MPLS Routes VRF's HW mode Table",
                                          "Cisco 7600 MPLS Routes/VRF's/HW mode Table") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("Hardware Table - VSS" in tableName):
                    if (tableName.replace("Hardware Table - VSS", "Catalyst 6500 Table - VSS") in item):
                        NMSArea = key
                        print(NMSArea)
                        break


                elif ("Catalyst 6500 and Cisco 7600 Series High Availability(HA)-System Audit" in tableName):
                    if (tableName.replace("Catalyst 6500 and Cisco 7600 Series High Availability(HA)-System Audit",
                                          "Catalyst 6500 and Cisco 7600 Series High Availability(HA) - System Audit") in item):
                        NMSArea = key
                        print(NMSArea)
                        break




                elif ("DR BDR Implementation - Main Domain" in tableName):
                    if (tableName.replace("DR BDR Implementation - Main Domain",
                                          "DR/BDR Implementation - Main Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("DR BDR Implementation - Other Domains" in tableName):
                    if (tableName.replace("DR BDR Implementation - Other Domains",
                                          "DR/BDR Implementation - Other Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break



                elif ("DR BDR Related (Exceptions only) - Main Domain" in tableName):
                    if (tableName.replace("DR BDR Related (Exceptions only) - Main Domain",
                                          "DR/BDR Related - Main Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("DR BDR Related (Exceptions only) - Other Domains" in tableName):
                    if (tableName.replace("DR BDR Related (Exceptions only) - Other Domains",
                                          "DR/BDR Related - Other Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("PE-CE RIP Version 2 Authentication(IOS, IOS XR and IOS XE)" in tableName):
                    if (tableName.replace("PE-CE RIP Version 2 Authentication(IOS, IOS XR and IOS XE)",
                                          "PE-CE RIP Version 2 Authentication") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("Audited ABRs in Main Domain" in tableName):
                    if (tableName.replace("Audited ABRs in Main Domain", "Audited ABRs - Main Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break


                elif ("Not Audited ABRs in Main Domain" in tableName):
                    if (tableName.replace("Not Audited ABRs in Main Domain", "Non Audited ABRs - Main Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("Largest Number of OSPF Routes per Type - Main Domain" in tableName):
                    NMSArea = "Configuration Management"
                    print(NMSArea)
                    break
                elif ("OSPF Memory Analysis for IOSXR" in tableName):
                    NMSArea = "Capacity Management"
                    print(NMSArea)
                    break

                elif ("OSPF SPF and LSA Throttle Timers" in tableName):
                    NMSArea = "Configuration Management"
                    print(NMSArea)
                    break

                elif ("BGP Generic Configuration Best Practices (IOS, IOS XR and IOS-XE)" in tableName):
                    if (tableName.replace("BGP Generic Configuration Best Practices (IOS, IOS XR and IOS-XE)",
                                          "BGP Generic Configuration best practices (IOS, IOS XR and IOS-XE)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break



                elif ("OSPF Neighbor Inventory (Exceptions only) - Main Domain" in tableName):
                    if (tableName.replace("OSPF Neighbor Inventory (Exceptions only) - Main Domain",
                                          "OSPF Neighbor Inventory - Main Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("OSPF Neighbor Inventory (Exceptions only) - Other Domains" in tableName):
                    if (tableName.replace("OSPF Neighbor Inventory (Exceptions only) - Other Domains",
                                          "OSPF Neighbor Inventory - Other Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("Neighbor related(Exceptions only)" in tableName):
                    if (tableName.replace("Neighbor related(Exceptions only)",
                                          "Neighbor related (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("IPv4 - Maximum Prefixes imported into VRF from CE EBGP peer (IOS and IOS XR)" in tableName):
                    if (tableName.replace(
                            "IPv4 - Maximum Prefixes imported into VRF from CE EBGP peer (IOS and IOS XR)",
                            "Maximum Prefixes imported into VRF from CE EBGP peer") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("BGP VRF Import Best Practices" in tableName):
                    if (tableName.replace("BGP VRF Import Best Practices", "BGP VRF Import best practices") in item):
                        NMSArea = key
                        print(NMSArea)
                        break


                elif ("Area 0 - Capacity - Addressing - Main Domain" in tableName):
                    if (tableName.replace("Area 0 - Capacity - Addressing - Main Domain",
                                          "Area 0 - Capacity - Addressing Table 1 - Main Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("Area 0 - Capacity - Addressing - Other Domains" in tableName):
                    if (tableName.replace("Area 0 - Capacity - Addressing - Other Domains",
                                          "Area 0 - Capacity - Addressing Table 1 - Other Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("Redistribution - External Route - ASBR Related (Exceptions only) - Main Domain" in tableName):
                    if (tableName.replace(
                            "Redistribution - External Route - ASBR Related (Exceptions only) - Main Domain",
                            "Redistribution - External Route - ASBR Related - Main Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("Redistribution - External Route - ASBR Related (Exceptions only) - Other Domain" in tableName):
                    if (tableName.replace(
                            "Redistribution - External Route - ASBR Related (Exceptions only) - Other Domain",
                            "Redistribution - External Route - ASBR Related - Other Domain (Exceptions only)") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                elif ("BGP Neighbor Performance and Convergence (IOS and IOS XR)" in tableName):
                    if (tableName.replace("BGP Neighbor Performance and Convergence (IOS and IOS XR)",
                                          "BGP Neighbor performance and convergence") in item):
                        NMSArea = key
                        print(NMSArea)
                        break

                elif ("IPv4 - Maximum Prefixes allowed in VRF (IOS ,IOS XR and IOS-XE)" in tableName):
                    NMSArea = "Performance Management"
                    print(NMSArea)
                    break

                if ("IPv4" in tableName):
                    if (tableName.replace("IPv4", "IPV4") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                if ("Domains" in tableName):
                    if (tableName.replace("Domains", "Domain") in item):
                        NMSArea = key
                        print(NMSArea)
                        break
                if (tableName + " Table" in item):
                    NMSArea = key
                    print(NMSArea)
                    break

                if (item.lower() in tableName.lower()):
                    NMSArea = key
                    print(NMSArea)
                    break

    return NMSArea


def createSeverityBreakdownJson(audit_id):
    severities = {"Verify Manually": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Configuration Management": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Capacity Management": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Fault Management": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Performance Management": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Security Management": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0},
                  "Total": {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0}}
    os.chdir(scriptPath + "/ExceptionTables" + audit_id + "/json")
    jsonData = []
    for file in os.listdir(scriptPath + "/ExceptionTables" + audit_id + "/json/"):
        if (file.endswith(".json")):
            with open(scriptPath + "/ExceptionTables" + audit_id + "/json/" + file, "r") as f:
                jsonData = json.loads(f.read())
            if (jsonData["NMSArea"]):
                severities["Total"]["Critical"] = severities["Total"]["Critical"] + jsonData["Critical"]
                severities["Total"]["High"] = severities["Total"]["High"] + jsonData["High"]
                severities["Total"]["Medium"] = severities["Total"]["Medium"] + jsonData["Medium"]
                severities["Total"]["Low"] = severities["Total"]["Low"] + jsonData["Low"]
                severities["Total"]["Informational"] = severities["Total"]["Informational"] + jsonData["Informational"]

                severities[jsonData["NMSArea"]]["Critical"] = severities[jsonData["NMSArea"]]["Critical"] + jsonData[
                    "Critical"]
                severities[jsonData["NMSArea"]]["High"] = severities[jsonData["NMSArea"]]["High"] + jsonData["High"]
                severities[jsonData["NMSArea"]]["Medium"] = severities[jsonData["NMSArea"]]["Medium"] + jsonData[
                    "Medium"]
                severities[jsonData["NMSArea"]]["Low"] = severities[jsonData["NMSArea"]]["Low"] + jsonData["Low"]
                severities[jsonData["NMSArea"]]["Informational"] = severities[jsonData["NMSArea"]]["Informational"] + \
                                                                   jsonData["Informational"]
            # else:
            # print(file)
    return (severities)


def createExceptionTable_AllDataTable(audit_id):
    ExceptionData = []
    fullData = []
    jsonData = []
    for file in os.listdir(scriptPath + "/ExceptionTables" + audit_id + "/json/"):
        if (file.endswith(".json")):
            with open(scriptPath + "/ExceptionTables" + audit_id + "/json/" + file, "r") as f:
                jsonData = json.loads(f.read())
            if (jsonData["NMSArea"] != "" and len(jsonData["data"]) > 1):
                for item in jsonData["data"]:
                    exceptionline = {}
                    fulldataline = {}
                    if (list(item.keys())[0] != "Host Name (IP Address)" and list(item.keys())[0] != "NREPs" and
                            list(item.keys())[0] != "No data available."):
                        # print(file)
                        # print(list(item.keys())[0])
                        item[list(item.keys())[0]]["Sev"] = item[list(item.keys())[0]]["Sev"].replace("/xa0", "")
                        if (item[list(item.keys())[0]]["Sev"] != "White"):
                            # print(item[list(item.keys())[0]]["Sev"])
                            exceptionline = {"Table Name": file.split(".")[0],
                                             "Host Name (IP Address)": item[list(item.keys())[0]]["device_name"],
                                             "Chassis Type": item[list(item.keys())[0]]["chassis_type"],
                                             "Severity": item[list(item.keys())[0]]["Sev"],
                                             "NMS Area": jsonData["NMSArea"],
                                             "Exception Name": list(item.keys())[0].strip(),
                                             "Exception Value": item[list(item.keys())[0]]["value"],
                                             "Row Number": item[list(item.keys())[0]]["row_number"]}
                        fulldataline = {"Table Name": file.split(".")[0],
                                        "Host Name (IP Address)": item[list(item.keys())[0]]["device_name"],
                                        "Chassis Type": item[list(item.keys())[0]]["chassis_type"],
                                        "NMS Area": jsonData["NMSArea"], "Column Name": list(item.keys())[0],
                                        "Column Value": item[list(item.keys())[0]]["value"],
                                        "Row Number": item[list(item.keys())[0]]["row_number"]}
                        if exceptionline:
                            ExceptionData.append(exceptionline)
                        fullData.append(fulldataline)
    print("Total Exceptions :: " + str(len(ExceptionData)))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/allExceptions.json", "w") as f:
        f.write(json.dumps(ExceptionData))
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/allData.json", "w") as f:
        f.write(json.dumps(fullData))


def createJson(file_name, jsonDir, audit_id):
    print("parsing file in createJson ::::: " + file_name)
    fName = getTableName(file_name)[0]
    fName = fName.strip()
    # if ("(Exceptions Only" in fName):
    # fName = fName.split("(Exceptions")[0][:-1]

    # print(" --- in CreateJSON :: File :: " +file_name)
    if (os.path.exists(jsonDir + "/" + fName + ".json")):
        with open(jsonDir + "/" + fName + ".json", "r") as f:
            jsonData = json.loads(f.read())
    else:
        jsonData = {"NMSArea": "", "Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0,
                    "columnNames": [], "data": []}

    print("Getting NMS Area for ::: " + fName)
    jsonData["NMSArea"] = getNMSArea(fName, audit_id)
    if ("/summaryTables" not in file_name):
        if (jsonData["NMSArea"] == ""):
            jsonData["NMSArea"] = "Verify Manually"
    print("-------------------------------------")
    column_names = getColumnNames(file_name)
    jsonData["columnNames"] = column_names

    table_data, sev_data = getTableData(file_name)
    print("-----------------------")
    print("For file : " + file_name)
    print(len(table_data))
    print(len(sev_data))
    print("-----------------------")
    if (len(table_data) % len(column_names) != 0):
        for filler in range(0, len(column_names) - (len(table_data) % len(column_names))):
            table_data.append("0")
            sev_data.append("White")

    table_data_list = []
    sev_data_list = []
    if (len(table_data) != 0):
        x = 0
        y = len(table_data)
        for i in range(x, y, len(column_names)):
            x = i
            table_data_list.append(table_data[x:x + len(column_names)])
            sev_data_list.append(sev_data[x:x + len(column_names)])

        for i in range(0, len(table_data_list)):
            if ("OSPF Neighbor Table for each sample (Exceptions" in fName):
                devName = table_data_list[i][1]
            else:
                devName = table_data_list[i][0]
            chassis = getChassisName(devName.split("\u00a0")[0], audit_id)
            if ("\u00a0" in devName):
                devName = devName.strip().split("\u00a0")[0] + " " + devName.strip().split("\u00a0")[-1]
            # print("number of columns in table are ::: "+str(len(column_names))+" and row number is "+ str(i))
            for j in range(0, len(column_names)):
                jjson = {}
                jjson[column_names[j]] = {"device_name": devName, "chassis_type": chassis,
                                          "value": table_data_list[i][j], "Sev": sev_data_list[i][j],
                                          "row_number": i + 1}
                jsonData["data"].append(jjson)
                if (sev_data_list[i][j] == "Low"):
                    jsonData["Low"] = jsonData["Low"] + 1
                elif (sev_data_list[i][j] == "High"):
                    jsonData["High"] = jsonData["High"] + 1
                elif (sev_data_list[i][j] == "Medium"):
                    jsonData["Medium"] = jsonData["Medium"] + 1
                elif (sev_data_list[i][j] == "Critical"):
                    jsonData["Critical"] = jsonData["Critical"] + 1
                elif (sev_data_list[i][j] == "Informational"):
                    jsonData["Informational"] = jsonData["Informational"] + 1

        with open(jsonDir + "/" + fName + ".json", "w") as jsonFile:
            jsonFile.write(json.dumps(jsonData))


def tableToNMSMapping(file_name, defaultJson):
    html_data = ""
    flag = 0
    print("In table to NMS mapping ::: " + file_name)
    table_to_NMS_Mapping_json = defaultJson
    with open(file_name, "r") as f:
        html_data = f.read()
    soup = BeautifulSoup(html_data, features="html.parser")
    anchors = soup.find_all("a")
    for anchor in anchors:
        if (
                "System Analysis" in anchor.text.strip() or "Media Analysis" in anchor.text.strip() or "Protocol Analysis" in anchor.text.strip()):
            html_data = html_data.replace("<b>Fault Management</b>",
                                          "<b><a NAME=Fault Management>Fault Management</a></b>")
            html_data = html_data.replace("<b>Capacity Management</b>",
                                          "<b><a NAME=Capacity Management>Capacity Management</a></b>")
            html_data = html_data.replace("<b>Configuration Management</b>",
                                          "<b><a NAME=Configuration Management>Configuration Management</a></b>")
            html_data = html_data.replace("<b>Performance Management</b>",
                                          "<b><a NAME=Performance Management>Performance Management</a></b>")
            html_data = html_data.replace("<b>Security Management</b>",
                                          "<b><a NAME=Security Management>Security Management</a></b>")
            break

    soup = BeautifulSoup(html_data, features="html.parser")
    anchors = soup.find_all("a")

    for anchor in anchors:
        print(anchor.text.strip())
        name = anchor.text.strip()
        if (name != ""):
            if ("Management" in name and "Table" not in name):
                key = name
                flag = 1
            else:
                if(flag == 1 and key == "Perormance Management"):
                    key = "Performance Management"
                if (flag == 1 and key in table_to_NMS_Mapping_json.keys()):
                    table_to_NMS_Mapping_json[key].append(name)
                elif (flag == 1):
                    table_to_NMS_Mapping_json[key] = [name]

    return (table_to_NMS_Mapping_json)


def convertTable(stat_table):
    # print("in convertTable : "+stat_table)
    tbl = str(stat_table)
    rowPtr = []
    tbl = tbl[tbl.find('</thead>') + 8:]
    tbl_list = tbl.split("</td>")
    for i in range(0, len(tbl_list) - 1):
        if ("<tr" in tbl_list[i]):
            currItem = tbl_list[i]
            tbl_list[i] = currItem[0: currItem.find(">") + 1]
            tbl_list.insert(i + 1, currItem[currItem.find(">") + 1:])

    for i in range(0, len(tbl_list)):
        if ("</tr><tr" in tbl_list[i]):
            currItem = tbl_list[i]
            tbl_list[i] = currItem[0:5]
            currItem = currItem[5:]
            tbl_list.insert(i + 1, currItem[0:currItem.find(">") + 1])
            tbl_list.insert(i + 2, currItem[currItem.find(">") + 1:])

    for i in range(0, len(tbl_list)):
        if ("</tr><tr" in tbl_list[i]):
            currItem = tbl_list[i]
            tbl_list[i] = currItem[0:5]
            currItem = currItem[5:]
            tbl_list.insert(i + 1, currItem[0:currItem.find(">") + 1])
            tbl_list.insert(i + 2, currItem[currItem.find(">") + 1:])

    for i in range(0, len(tbl_list)):
        if ("</tr><tr" in tbl_list[i]):
            currItem = tbl_list[i]
            tbl_list[i] = currItem[0:5]
            currItem = currItem[5:]
            tbl_list.insert(i + 1, currItem[0:currItem.find(">") + 1])
            tbl_list.insert(i + 2, currItem[currItem.find(">") + 1:])
    for i in range(0, len(tbl_list)):
        if ("</tr><tr" in tbl_list[i]):
            currItem = tbl_list[i]
            tbl_list[i] = currItem[0:5]
            currItem = currItem[5:]
            tbl_list.insert(i + 1, currItem[0:currItem.find(">") + 1])
            tbl_list.insert(i + 2, currItem[currItem.find(">") + 1:])

    # print(tbl_list)
    for i in range(0, len(tbl_list) - 1):
        if ("<tr" in tbl_list[i]):
            rowPtr.append(i)
    w, h = (rowPtr[1] - rowPtr[0]), len(rowPtr)
    Matrix = [[0 for x in range(w)] for y in range(h)]
    j, k = -1, 0
    for i in range(0, len(tbl_list)):

        if ("<tr" in tbl_list[i]):
            j = j + 1
            k = 0
        while (Matrix[j][k] != 0):
            k = k + 1

        if ("tr" not in tbl_list[i]):
            rowspan = re.findall("rowspan=\"\d\"", tbl_list[i])
            if (not rowspan):
                rowspan = ["rowspan=1 "]
            # print(int(rowspan[0][-2:-1]))
            if (int(rowspan[0][-2:-1]) > 1):
                for locPtr in range(1, int(rowspan[0][-2:-1])):
                    # print(j+locPtr , k)
                    # print(re.sub("rowspan=\"\d\"" , '' , tbl_list[i]))
                    Matrix[j + locPtr][k] = re.sub("rowspan=\"\d\"", '', tbl_list[i])

        if ("tr" not in tbl_list[i]):
            Matrix[j][k] = re.sub("rowspan=\"\d\"", '', tbl_list[i]) + "</td>"
        else:
            Matrix[j][k] = re.sub("rowspan=\"\d\"", '', tbl_list[i])
        # k = k + 1

    tableStr = str(stat_table)
    tableStr = tableStr[0: tableStr.find('</thead>') + 8]
    # print(tbl_list)
    for item in Matrix:
        for k in range(0, len(item)):
            if (item[k] != 0):
                tableStr = tableStr + item[k]

    return (tableStr)


def getTableData(file_name):
    html_data = ""
    with open(file_name, "r") as f:
        html_data = f.read()
    html_data = html_data.replace("<B>", "")
    html_data = html_data.replace("</B>", "")
    soup = BeautifulSoup(html_data, features="html.parser")
    stat_table = soup.find('table')
    # print(stat_table)
    # stat_table=stat_table[0]
    data = []
    sev_data = []
    rowspanflag = 0
    doneFlag = 0
    for row in stat_table.find_all('tr'):
        for cell in row.find_all('td', class_=True):
            if ("rowspan" in str(cell.encode("utf-8"))):
                stat_table = convertTable(stat_table)
                # print(stat_table)
                doneFlag = 1
            if (doneFlag == 1):
                break
        if (doneFlag == 1):
            break

    if (isinstance(stat_table, str)):
        soup = BeautifulSoup(stat_table, features="html.parser")
        stat_table = soup.find('table')

    for row in stat_table.find_all('tr'):
        for cell in row.find_all('td', class_=True):
            y = cell.text.replace("\xa0", " ")
            sev = cell["class"][0]
            if 'ADDFFF' in sev:
                sev_data.append("Critical")
            elif 'FF0000' in sev:
                sev_data.append("High")
            elif 'F9966B' in sev:
                sev_data.append("Medium")
            elif 'FFFF00' in sev:
                sev_data.append("Low")
            elif '00FF00' in sev:
                sev_data.append("Informational")
            else:
                sev_data.append("White")

            data.append(y)

    return (data, sev_data)


def getTableName(file_name):
    html_data = ""
    tableName = []
    with open(file_name, "r", encoding="utf-8") as f:
        html_data = f.read()
    if (len(html_data) > 2):
        soup = BeautifulSoup(html_data, features="html.parser")
        x = soup.find("thead")
        # print("in getTableName :: " + str(x))
        if (x):
            rows = x.findChildren('tr')
            tableName = [x.getText() for x in rows[0].findChildren('td')]
            tableName[0] = tableName[0].replace("/", " ")[0:100]
            return (tableName)
        else:
            return (file_name.split(".")[0])


def getColumnNames(file_name):
    html_data = ""
    columnNames = []
    with open(file_name, "r") as f:
        html_data = f.read()

    soup = BeautifulSoup(html_data, features="html.parser")
    x = soup.find("thead")
    rows = x.findChildren('tr')
    print("Heading Rows : ", len(rows))
    if (len(rows) > 1):
        rows = rows[1:]
    k = 0
    for row in reversed(rows):
        k = k + 1
        currentPointer = 0
        cols = row.findChildren('td')

        for col in cols:
            if (int(col["colspan"]) == 1 and int(col["rowspan"]) == k):
                columnNames.insert(currentPointer, col.getText().replace("\xa0", ""))
                currentPointer = currentPointer + 1
            else:
                for i in range(0, int(col["colspan"])):
                    # print(i)
                    temp = columnNames[currentPointer + i]
                    temp = col.getText() + " - " + temp
                    columnNames[currentPointer + i] = temp
                currentPointer = currentPointer + int(col["colspan"])

    return (columnNames)


def createAuditInfo(html_data, audit_id):
    print(" ::: In create audit info :::")
    audit_info_json = {}
    soup = BeautifulSoup(html_data, features="html.parser")
    for row in soup.select('tr'):
        row_text = [x.text for x in row.find_all('td')]
        audit_info_json[row_text[0]] = row_text[1]
    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/audit_information.json", "w") as f:
        f.write(json.dumps(audit_info_json))
    print(" ::: In create audit info :::")
    print(audit_info_json)


def getAuditSummary(file_name, audit_id):
    with open(file_name, "r", encoding="utf-8") as f:
        html_data = str(f.read().replace("0xc3", ""))
    print("in getAuditSummary :: " + file_name)
    soup = BeautifulSoup(html_data, features="html.parser")
    tables = soup.find_all("table")

    for table in tables:
        head = table.find_all("thead")
        for h in head:
            rows = h.findChildren('tr')
            tableName = [x.getText() for x in rows[0].findChildren('td')]
            if (tableName[0] != ""):
                if (len(tableName) == 1):
                    tableName[0] = tableName[0].replace("/", " ")[0:100]
                    with open("summaryTables/" + tableName[0] + ".html", "w") as f:
                        tableData1 = str(table.encode("utf-8")).replace("/xa0", " ")
                        f.write(tableData1)
                elif ("Exception Type" in tableName[0]):
                    with open("summaryTables/" + "Overall Exceptions.html", "w") as f:
                        tableData1 = str(table.encode("utf-8")).replace("/xa0", " ")
                        f.write(tableData1)
            else:
                print("No table name")
                global headlesstableFlag
                if (headlesstableFlag == 0):
                    if (len(tableName) == 1 and "Audit_Summary.html" in file_name):
                        createAuditInfo(
                            "<table><thead></thead>" + str(table).split("</thead>")[1][:-8] + "</tbody></table>",audit_id)
                        headlesstableFlag = 1


def createNetRuleNetAdvices(filePath, audit_id):
    jsonDataList = []
    wb = openpyxl.load_workbook(filePath + "/" + 'Audit_Summary_Excel.xlsx')
    active_sheet = wb["Audit_Exception"]
    for i in range(5, active_sheet.max_row + 1):

        exception = active_sheet.cell(row=i, column=1).value
        exception = exception[exception.rfind(',"') + 2:-2]
        if ("<br/>" in exception):
            exception = exception.replace("<br/>", "")
        table_name = active_sheet.cell(row=i, column=2).value
        net_rule = active_sheet.cell(row=i, column=7).value
        net_advice = active_sheet.cell(row=i, column=8).value
        tempJson = {"exception": exception, "table_name": table_name, "net_rule": net_rule, "net_advice": net_advice}
        jsonDataList.append(tempJson)

    with open(scriptPath + "/ExceptionTables" + audit_id + "/data/netRulenetAdvice.json", "w") as f:
        f.write(json.dumps(jsonDataList))


def Createfiles(tdict, fileName, audit_id):
    os.chdir(scriptPath + "/ExceptionTables" + audit_id)
    if not os.path.exists("Add_Remaining"):
        os.makedirs("Add_Remaining")

    for tablename, folders in tdict.items():

        for folder in folders:
            for f in os.listdir(scriptPath + "/" + fileName + "/Full/" + folder):
                if f.endswith(".html"):

                    with open(scriptPath + "/" + fileName + "/Full/" + folder + "/" + f, 'r') as F:
                        html_data = F.read()

                        soup = BeautifulSoup(html_data, features="html.parser")
                        tables = soup.find_all("table")

                        for table in tables:
                            if tablename in table.getText():
                                tablename = tablename.replace("/", "")
                                with open(
                                        scriptPath + "/ExceptionTables" + audit_id + "/Add_Remaining/" + folder + '_' + tablename + '.html',
                                        'w') as W:
                                    tableData = str(table.encode("utf-8")).replace("/xa0", " ")
                                    W.write(tableData)


def addremaining(filePath, fileName, audit_id):
    Tabledict = {}
    TableList = []

    for f in os.listdir(scriptPath + "/ExceptionTables" + audit_id + "/json"):
        TableList.append(f.split('.')[0])

    print(TableList)

    wb = openpyxl.load_workbook(filePath + "/" + 'Audit_Summary_Excel.xlsx')
    active_sheet = wb["Audit_Exception"]

    for i in range(5, active_sheet.max_row + 1):

        table_name = active_sheet.cell(row=i, column=2).value

        if table_name not in TableList:
            nodes_sheet = wb[f'Exception Row-{i} nodes']

            for j in range(4, nodes_sheet.max_row + 1):
                device = nodes_sheet.cell(row=j, column=1).value

                device_dir = device.split('), "')
                # deviceName = device_dir[1].split(',')[1][1:-2]
                deviceFolder = device_dir[1].split('/')[0]

                if table_name in Tabledict:
                    Tabledict[table_name].append(deviceFolder)
                else:
                    Tabledict[table_name] = [deviceFolder]

    Createfiles(Tabledict, fileName, audit_id)

def send_mail(audit_id, mail_id):
    body1 = """Hi, 
            Audit processing initiated for audit ID """+str(audit_id)+""". Please allow upto 12 hours before raising 
            issue with bgl-cx-bcs-audit@cisco.com

            Regards,
            BCS Audit team """

    mail_utility.send_email(str(mail_id), 'Please do not reply to this mail.',body1,'')

def mainFunction(f_name):
    startTime = getTime("date")

    print("Script Initiating at : ", str(datetime.datetime.now()).split(".")[0])
    infoJson = []
    with open(f_name, "r") as f:
        infoJson = json.loads(f.read())
    fileName = infoJson["file_name"]

    send_mail(infoJson["audit_id"], infoJson["emp_cec_id"])

    print("Creating ExceptionTables directory. This will store all exception tables only")
    if not os.path.exists("ExceptionTables" + infoJson["audit_id"]):
        os.makedirs("ExceptionTables" + infoJson["audit_id"])
    extractZip(fileName)
    fileName = fileName.split(".")[0]

    # device name and model details
    devToChassisMap = device_to_chassis_mapping(scriptPath, fileName)
    os.chdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"])
    if not os.path.exists("data"):
        os.makedirs("data")
    if not os.path.exists("summaryTables"):
        os.makedirs("summaryTables")
    with open("data/" + "deviceDetails.json", "w") as f:
        f.write(json.dumps(devToChassisMap))

    defaultJson = {"Performance Management": [], "Capacity Management": [], "Fault Management": [],
                   "Configuration Management": [], "Security Management": []}
    print("Initiating data extraction from HTML tables")
    count = 1

    for item in os.listdir(scriptPath + "/" + fileName + "/Full/summary"):
        if ("Audit_Overview.html" in item):
            getAuditSummary(scriptPath + "/" + fileName + "/Full/summary/" + item, infoJson["audit_id"])
        elif ("Audit_Summary.html" in item):
            getAuditSummary(scriptPath + "/" + fileName + "/Full/summary/" + item, infoJson["audit_id"])
        elif ((item.endswith(".html") and "NodeAndSummary_" in item) or "Detailed_Findings.html" in item):
            tableToNMSMap = tableToNMSMapping(scriptPath + "/" + fileName + "/Full/summary/" + item, defaultJson)
            if (
            os.path.exists(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/data/" + "tableToNMSMap.json")):
                with open(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/data/" + "tableToNMSMap.json",
                          "r") as f:
                    jsonData = json.loads(f.read())
                for key in tableToNMSMap.keys():
                    for item in tableToNMSMap[key]:
                        jsonData[key].append(item)
                with open(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/data/" + "tableToNMSMap.json",
                          "w") as f:
                    f.write(json.dumps(jsonData))
            else:
                with open(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/data/" + "tableToNMSMap.json",
                          "w") as f:
                    f.write(json.dumps(tableToNMSMap))

        elif (os.path.isdir(scriptPath + "/" + fileName + "/Full/summary/" + item) and "chart" not in item):
            for f in os.listdir(scriptPath + "/" + fileName + "/Full/summary/" + item):
                if (".DS" not in f):
                    # print("item for extractTable is ::" + f)
                    tableData = extractTable(scriptPath + "/" + fileName + "/Full/summary/" + item + "/" + f)
                    if (len(tableData) > 1):
                        html_list = os.listdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"])
                        if (f in html_list):
                            file_to_save = f.split(".")[0] + "_" + str(count) + ".html"
                            with open(file_to_save, "w") as file:
                                file.write(tableData)
                                count = count + 1
                        else:
                            with open(f, "w") as file:
                                file.write(tableData)

    print("Creating JSON directory under ExceptionTables. This directory will store HTML table data in JSON format")

    os.chdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"])
    if not os.path.exists("json"):
        os.makedirs("json")
    jsonDir = scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/json"
    print("Creating JSON ... ")

    for f in os.listdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"]):
        if (f.endswith(".html")):
            createJson(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/" + f, jsonDir, infoJson["audit_id"])

    for f in os.listdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/summaryTables"):
        # print(scriptPath+"/ExceptionTables"+"/summaryTables"+f)
        if (f.endswith(".html")):
            createJson(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/summaryTables/" + f, jsonDir,
                       infoJson["audit_id"])

    addremaining(scriptPath + "/" + fileName + "/Full/summary", fileName, infoJson["audit_id"])
    for f in os.listdir(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/Add_Remaining"):
        if (f.endswith(".html")):
            createJson(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/Add_Remaining/" + f, jsonDir,
                       infoJson["audit_id"])

    severitiesJson = createSeverityBreakdownJson(infoJson["audit_id"])
    with open(scriptPath + "/ExceptionTables" + infoJson["audit_id"] + "/data/severityBreakdown.json", "w")as f:
        f.write(json.dumps(severitiesJson))

    # Creating all data and all exceptions json file
    print("Creating all Exceptions and All data JSON Now...")

    createExceptionTable_AllDataTable(infoJson["audit_id"])
    if (infoJson["audit_type"] == "Nexus Audit"):
        createSoftwareTable_Graphs(infoJson["audit_id"])
        createCPUandMemoryTable(infoJson["audit_id"])
    elif (infoJson["audit_type"] == "IOS XR Audit"):
        createIOSXRSoftwareHardwareGraph(infoJson["audit_id"])

    # create netRule netAdvice file
    createNetRuleNetAdvices(scriptPath + "/" + fileName + "/Full/summary/", infoJson["audit_id"])

    insertIntoDB(infoJson["customer_key"], infoJson["audit_id"], infoJson["customer_name"], infoJson["audit_type"])

    cna_graph_images.GraphInsertion(fileName, infoJson["customer_key"], infoJson["audit_id"], infoJson["customer_name"],
                                    infoJson["audit_type"])

    # print("Copying CSS file ...")
    # shutil.copy2(scriptPath+"/"+fileName+"/Full/commonImages/"+"reportStylesheet.css"  ,scriptPath+"/ExceptionTables/"+"reportStylesheet.css")

    # clean up code
    print("Cleaning up ...")
    if (os.path.exists(scriptPath + "/AuditReportViewer.htm")):
        os.remove(scriptPath + "/AuditReportViewer.htm")
    shutil.rmtree(scriptPath + "/" + fileName)
    os.remove(scriptPath + "/" + fileName + ".zip")
    os.remove(scriptPath + "/" + f_name)
    shutil.rmtree(scriptPath + "/ExceptionTables"+infoJson["audit_id"])
    # If this is not done, flag is remaining 1 and next time when audit is initiated, audit_info file is not getting created
    global headlesstableFlag
    headlesstableFlag = 0
    endTime = getTime("date")
    print("Time taken for exceution : ", str(endTime - startTime).split(".")[0])